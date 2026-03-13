"""
Excel to SQLite Adapter for TurboVault Engine.

This module provides the ExcelImport class, which converts an Excel file
into an in-memory SQLite database and then delegates the metadata import
to the SqliteImportService.
"""

from __future__ import annotations

import logging
import sqlite3
from typing import TYPE_CHECKING

import openpyxl

from engine.services.sqlite_import import SqliteImportService

if TYPE_CHECKING:
    from engine.models.project import Project

logger = logging.getLogger(__name__)


class ExcelImport:
    """
    Adapter to import metadata from the legacy Excel format.

    Uses openpyxl to read the workbook, converts it into an in-memory SQLite
    database, and then uses SqliteImportService to perform the actual import.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        self._conn = sqlite3.connect(":memory:")
        self._conn.row_factory = sqlite3.Row

        # Fill the SQLite DB with sheet data
        for sheet_name in self._wb.sheetnames:
            self._load_sheet_to_sqlite(sheet_name)

    def _load_sheet_to_sqlite(self, sheet_name: str) -> None:
        """Read an openpyxl worksheet and insert all rows into a SQLite table."""
        ws = self._wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            return  # empty sheet

        columns = [str(c).strip().lower() for c in header_raw if c is not None]
        if not columns:
            return

        # Sanitise table name (sheet names may contain spaces)
        tbl = sheet_name  # we quote it in SQL with []

        col_defs = ", ".join(f"[{c}] TEXT" for c in columns)
        self._conn.execute(f"CREATE TABLE IF NOT EXISTS [{tbl}] ({col_defs})")

        placeholders = ", ".join("?" for _ in columns)
        insert_sql = f"INSERT INTO [{tbl}] ({', '.join(f'[{c}]' for c in columns)}) VALUES ({placeholders})"

        num_cols = len(columns)
        for row in rows_iter:
            # Trim row to header length
            vals = list(row[:num_cols])
            # Convert non-None values to strings
            vals = [str(v).strip() if v is not None else None for v in vals]
            # Replace stringified null markers with actual None
            vals = [
                None if v is not None and v.strip().lower() in ("none", "") else v
                for v in vals
            ]
            self._conn.execute(insert_sql, vals)

        self._conn.commit()

    def import_metadata(
        self,
        project_name: str | None = None,
        description: str | None = None,
        project: Project | None = None,
        skip_snapshots: bool = False,
    ) -> Project:
        """
        Adapts the Excel data and imports it using SqliteImportService.
        """
        service = SqliteImportService(self._conn)
        try:
            return service.import_metadata(
                project_name=project_name,
                description=description or f"Imported from Excel: {self.file_path}",
                project=project,
                skip_snapshots=skip_snapshots,
            )
        finally:
            self.close()

    def close(self) -> None:
        """Close the in-memory SQLite connection and workbook."""
        if self._conn:
            self._conn.close()
        if self._wb:
            self._wb.close()
            self._wb = None
