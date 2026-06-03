"""IRiS parser: the three IRiS Excel files -> DomainModel.

Inverse of ``engine.services.export.exporters.iris_exporter``. Reads the
``Source_*.xlsx`` / ``DataVault_*.xlsx`` / ``Mappings_*.xlsx`` triple and builds
a ``DomainModel`` directly, bypassing the IR / schema-validation layer (the same
way ``parse_json`` does), so the rest of the import pipeline (plan -> execute)
and dbt generation work unchanged.

The reconstruction reverses the conventions the exporter applies:

  - ``BKCC`` / ``BKCC_<HUB>`` collision-code columns (and their Source columns
    and Mappings rows) are excluded.
  - The payload satellite the exporter writes for a non-historized link is
    folded back into the link as payload columns, and the link is marked
    ``non_historized``.
  - The secondary hub mappings that relate a link's business key to the
    connected hub's key are used to rebuild the link's hub-source mappings.

IRiS represents hubs, links, satellites and their column mappings. Metadata
outside that scope is not part of the IRiS files and so is not present in the
result: reference hubs/satellites, effectivity / record-tracking satellites,
prejoins, hashkey and group names, source-system identity beyond the schema,
``is_primary_source``, and the ``_H``/``_L``/``_S`` name suffixes (reconstructed
physical names are the IRiS base names).
"""

from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from engine.services.export.exporters.iris_exporter import (
    IRIS_COL_BKCC,
    IRIS_COL_BUSINESS_KEY,
    IRIS_COL_LINK_BKCC,
    IRIS_COL_LINK_BUSINESS_KEY,
    IRIS_SUBTYPE_MULTI_ACTIVE,
    IRIS_TABLE_TYPE_HUB,
    IRIS_TABLE_TYPE_LINK,
    IRIS_TABLE_TYPE_SATELLITE,
)
from engine.services.imports.domain import (
    DHub,
    DHubColumn,
    DHubSourceMapping,
    DLink,
    DLinkColumn,
    DLinkHubReference,
    DLinkHubSourceMapping,
    DLinkSourceMapping,
    DomainModel,
    DSatellite,
    DSatelliteColumn,
    DSourceColumn,
    DSourceSystem,
    DSourceTable,
)
from engine.services.imports.errors import Code, PipelineAbort, make_issue
from engine.services.imports.parsers.base import clean
from engine.services.imports.parsers.excel import _read_sheet
from engine.services.imports.ir import IRRow
from engine.services.imports.types import IssueLocation

# Each IRiS file is matched by a token in its name. The match is case- and
# separator-insensitive (the name is lower-cased with non-alphanumerics removed),
# so "Source_Template", "template source", "DataVault", "data vault" and
# "col-mappings" all match.
_SOURCE_TOKENS = ("source",)
_DATAVAULT_TOKENS = ("datavault", "vault")
_MAPPINGS_TOKENS = ("mapping",)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def parse_iris(path: Path) -> DomainModel:
    """Read an IRiS three-file export from a directory and build a DomainModel.

    ``path`` is a directory holding the three workbooks, matched by a ``source``
    / ``datavault`` / ``mapping`` token anywhere in the file name (case- and
    separator-insensitive). Raises ``PipelineAbort`` if the directory or any
    of the three files cannot be found or read.
    """
    if not path.exists():
        raise PipelineAbort(_abort_issue(f"Path not found: {path}", path))
    if not path.is_dir():
        raise PipelineAbort(
            _abort_issue(
                f"IRiS import expects a directory of three files, got a file: {path}",
                path,
                suggestion="Point --source at the folder containing the "
                "Source, DataVault and Mappings workbooks.",
            )
        )

    xlsx_files = [
        p for p in path.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"
    ]
    source_file = _find_file(xlsx_files, _SOURCE_TOKENS)
    datavault_file = _find_file(xlsx_files, _DATAVAULT_TOKENS)
    mappings_file = _find_file(xlsx_files, _MAPPINGS_TOKENS)

    missing = [
        label
        for label, found in (
            ("Source", source_file),
            ("DataVault", datavault_file),
            ("Mappings", mappings_file),
        )
        if found is None
    ]
    if missing:
        raise PipelineAbort(
            _abort_issue(
                f"IRiS directory {path} is missing required file(s): "
                f"{', '.join(missing)}.",
                path,
                suggestion="An IRiS import needs all three workbooks "
                "(Source, DataVault and Mappings) in the same folder.",
            )
        )

    source_rows = _load_first_sheet(source_file)
    datavault_rows = _load_first_sheet(datavault_file)
    mappings_rows = _load_first_sheet(mappings_file)

    return _Resolver(source_rows, datavault_rows, mappings_rows).build()


# ---------------------------------------------------------------------------
# File discovery / reading
# ---------------------------------------------------------------------------


def _find_file(xlsx_files: list[Path], tokens: tuple[str, ...]) -> Path | None:
    """Return the first ``.xlsx`` whose normalized name contains one of
    ``tokens``. The name is lower-cased with non-alphanumerics removed, so
    casing and separators do not matter."""
    matches = sorted(
        p
        for p in xlsx_files
        if any(token in re.sub(r"[^a-z0-9]", "", p.stem.lower()) for token in tokens)
    )
    return matches[0] if matches else None


def _load_first_sheet(path: Path) -> list[IRRow]:
    """Read the first sheet of an IRiS workbook into cleaned rows.

    IRiS data always lives on the first sheet (``Sheet1``); the DataVault file's
    second sheet is just an enum reference and is ignored.
    """
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError) as exc:
        raise PipelineAbort(
            _abort_issue(
                f"Could not open IRiS file {path.name}: {exc}",
                path,
                suggestion="Confirm the file is a valid .xlsx workbook.",
            )
        ) from exc

    try:
        if not wb.sheetnames:
            raise PipelineAbort(
                make_issue(
                    severity="error",
                    code=Code.SOURCE_EMPTY,
                    stage="parse",
                    message=f"IRiS file {path.name} contains no sheets.",
                    location=IssueLocation(file=str(path)),
                )
            )
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
        sheet = _read_sheet(ws, ws.title)
    finally:
        wb.close()

    return list(sheet.rows) if sheet is not None else []


def _abort_issue(
    message: str, path: Path, *, suggestion: str | None = None
):
    return make_issue(
        severity="error",
        code=Code.SOURCE_UNREADABLE,
        stage="parse",
        message=message,
        location=IssueLocation(file=str(path)),
        suggestion=suggestion,
    )


# ---------------------------------------------------------------------------
# Small value helpers
# ---------------------------------------------------------------------------


def _is_bkcc_column(name: str | None) -> bool:
    """True for the synthesized BKCC / BKCC_<HUB> collision-code columns."""
    if not name:
        return False
    upper = name.upper()
    return upper == IRIS_COL_BKCC.upper() or upper.startswith("BKCC_")


def _as_int(value: object) -> int | None:
    if value is None:
        return None
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return None


def _compose_datatype(
    base: str | None, size: object, scale: object
) -> str:
    """Recompose a TurboVault datatype string from IRiS's separate cells, e.g.
    (``varchar``, 4000, None) -> ``varchar(4000)`` and
    (``decimal``, 18, 4) -> ``decimal(18,4)``."""
    base = (base or "").strip()
    if not base:
        return ""
    n_size = _as_int(size)
    if n_size is None:
        return base
    n_scale = _as_int(scale)
    if n_scale is None:
        return f"{base}({n_size})"
    return f"{base}({n_size},{n_scale})"


def _strip_prefix(name: str) -> tuple[str, str | None, bool]:
    """Split an IRiS entity name into ``(base, kind, is_multi_active)``.

    ``kind`` is ``hub`` / ``link`` / ``satellite`` / ``None`` (unprefixed).
    """
    low = name.lower()
    if low.startswith("s_ma_"):
        return name[5:], "satellite", True
    if low.startswith("s_"):
        return name[2:], "satellite", False
    if low.startswith("h_"):
        return name[2:], "hub", False
    if low.startswith("l_"):
        return name[2:], "link", False
    return name, None, False


# ---------------------------------------------------------------------------
# Working structures, keyed by the prefixed IRiS name.
# ---------------------------------------------------------------------------


class _HubWork:
    __slots__ = ("iris_name", "base", "bk_columns")

    def __init__(self, iris_name: str, base: str) -> None:
        self.iris_name = iris_name
        self.base = base
        self.bk_columns: list[str] = []


class _LinkWork:
    __slots__ = ("iris_name", "base", "hub_ref_iris", "additional_columns")

    def __init__(self, iris_name: str, base: str) -> None:
        self.iris_name = iris_name
        self.base = base
        # Related-hub IRiS names, ordered by first appearance, de-duplicated.
        self.hub_ref_iris: list[str] = []
        self.additional_columns: list[str] = []

    def add_hub_ref(self, hub_iris_name: str) -> None:
        if not hub_iris_name:
            return
        if any(hub_iris_name.lower() == existing.lower() for existing in self.hub_ref_iris):
            return
        self.hub_ref_iris.append(hub_iris_name)


class _SatWork:
    __slots__ = ("iris_name", "base", "is_multi_active", "parent_iris", "target_columns")

    def __init__(
        self, iris_name: str, base: str, is_multi_active: bool, parent_iris: str
    ) -> None:
        self.iris_name = iris_name
        self.base = base
        self.is_multi_active = is_multi_active
        self.parent_iris = parent_iris
        self.target_columns: list[str] = []


# ---------------------------------------------------------------------------
# Resolver
# ---------------------------------------------------------------------------


class _Resolver:
    def __init__(
        self,
        source_rows: list[IRRow],
        datavault_rows: list[IRRow],
        mappings_rows: list[IRRow],
    ) -> None:
        self._source_rows = source_rows
        self._datavault_rows = datavault_rows
        self._mappings_rows = mappings_rows

        self._model = DomainModel()
        # Mappings indexed by lower-cased target table -> list of
        # (source_table, source_column, target_column).
        self._maps_by_target: dict[str, list[tuple[str, str, str]]] = {}

    # -- public ---------------------------------------------------------- #

    def build(self) -> DomainModel:
        self._index_mappings()
        self._build_sources()
        hubs, links, sats = self._parse_datavault()
        hub_canon = self._assemble_hubs(hubs)
        link_by_iris, link_canon = self._assemble_links(links, hub_canon)
        self._assemble_satellites(sats, hub_canon, link_canon, link_by_iris)
        return self._model

    # -- mappings index -------------------------------------------------- #

    def _index_mappings(self) -> None:
        for row in self._mappings_rows:
            v = row.values
            src_table = clean(v.get("source table"))
            src_col = clean(v.get("source column"))
            tgt_table = clean(v.get("target table"))
            tgt_col = clean(v.get("target column"))
            if not (src_table and src_col and tgt_table and tgt_col):
                continue
            if _is_bkcc_column(tgt_col):
                continue
            self._maps_by_target.setdefault(tgt_table.lower(), []).append(
                (src_table, src_col, tgt_col)
            )

    def _target_mappings(self, iris_name: str) -> list[tuple[str, str, str]]:
        return self._maps_by_target.get(iris_name.lower(), [])

    # -- source file ----------------------------------------------------- #

    def _build_sources(self) -> None:
        tables_by_id: dict[str, DSourceTable] = {}
        for row in self._source_rows:
            v = row.values
            schema = clean(v.get("table schema")) or "source"
            table_name = clean(v.get("table name"))
            column = clean(v.get("column"))
            if not table_name or not column:
                continue
            if _is_bkcc_column(column):
                continue
            table = tables_by_id.get(table_name)
            if table is None:
                table = DSourceTable(
                    identifier=table_name, physical_name=table_name
                )
                tables_by_id[table_name] = table
                system = self._model.source_systems.setdefault(
                    schema, DSourceSystem(name=schema, schema_name=schema)
                )
                system.tables[table_name] = table
            datatype = _compose_datatype(
                clean(v.get("datatype")), v.get("size"), v.get("scale")
            )
            table.columns[column.lower()] = DSourceColumn(
                name=column, datatype=datatype
            )

    # -- datavault file -------------------------------------------------- #

    def _parse_datavault(
        self,
    ) -> tuple[dict[str, _HubWork], dict[str, _LinkWork], dict[str, _SatWork]]:
        hubs: dict[str, _HubWork] = {}
        links: dict[str, _LinkWork] = {}
        sats: dict[str, _SatWork] = {}

        for row in self._datavault_rows:
            v = row.values
            table_type = clean(v.get("table type"))
            iris_name = clean(v.get("table name"))
            column = clean(v.get("column"))
            if not table_type or not iris_name or not column:
                continue
            col_type = clean(v.get("column types"))

            if table_type == IRIS_TABLE_TYPE_HUB:
                self._parse_hub_row(hubs, iris_name, column, col_type)
            elif table_type == IRIS_TABLE_TYPE_LINK:
                self._parse_link_row(links, iris_name, column, col_type, v)
            elif table_type == IRIS_TABLE_TYPE_SATELLITE:
                self._parse_satellite_row(sats, iris_name, column, v)

        return hubs, links, sats

    def _parse_hub_row(
        self,
        hubs: dict[str, _HubWork],
        iris_name: str,
        column: str,
        col_type: str | None,
    ) -> None:
        if _is_bkcc_column(column) or col_type == IRIS_COL_BKCC:
            return
        if col_type != IRIS_COL_BUSINESS_KEY:
            return
        base = _strip_prefix(iris_name)[0]
        hub = hubs.setdefault(iris_name.lower(), _HubWork(iris_name, base))
        if not any(column.lower() == c.lower() for c in hub.bk_columns):
            hub.bk_columns.append(column)

    def _parse_link_row(
        self,
        links: dict[str, _LinkWork],
        iris_name: str,
        column: str,
        col_type: str | None,
        values: dict,
    ) -> None:
        base = _strip_prefix(iris_name)[0]
        link = links.setdefault(iris_name.lower(), _LinkWork(iris_name, base))
        # "Relationship" holds the related hub's prefixed IRiS name.
        related_hub = clean(values.get("relationship"))
        if col_type in (IRIS_COL_LINK_BUSINESS_KEY, IRIS_COL_LINK_BKCC):
            if related_hub:
                link.add_hub_ref(related_hub)
            return
        if _is_bkcc_column(column):
            return
        if not any(column.lower() == c.lower() for c in link.additional_columns):
            link.additional_columns.append(column)

    def _parse_satellite_row(
        self,
        sats: dict[str, _SatWork],
        iris_name: str,
        column: str,
        values: dict,
    ) -> None:
        base, _kind, is_ma_prefix = _strip_prefix(iris_name)
        subtype = clean(values.get("subtype"))
        parent_iris = clean(values.get("parent table")) or ""
        key = iris_name.lower()
        sat = sats.get(key)
        if sat is None:
            is_multi_active = is_ma_prefix or subtype == IRIS_SUBTYPE_MULTI_ACTIVE
            sat = _SatWork(iris_name, base, is_multi_active, parent_iris)
            sats[key] = sat
        if not any(column.lower() == c.lower() for c in sat.target_columns):
            sat.target_columns.append(column)

    # -- assembly: hubs -------------------------------------------------- #

    def _assemble_hubs(self, hubs: dict[str, _HubWork]) -> dict[str, str]:
        """Build hubs; return a case-insensitive map {lower(base) -> base} used
        to resolve hub references and satellite parents regardless of casing."""
        hub_canon: dict[str, str] = {}
        for hub in hubs.values():
            dhub = DHub(physical_name=hub.base, hub_type="standard")
            col_by_lower: dict[str, DHubColumn] = {}
            for i, bk in enumerate(hub.bk_columns):
                col = DHubColumn(
                    name=bk, column_type="business_key", sort_order=i + 1
                )
                dhub.columns.append(col)
                col_by_lower[bk.lower()] = col

            seen: set[tuple[str, str]] = set()
            for src_table, src_col, tgt_col in self._target_mappings(hub.iris_name):
                col = col_by_lower.get(tgt_col.lower())
                if col is None:
                    continue
                key = (src_table.lower(), src_col.lower())
                if key in seen:
                    continue
                seen.add(key)
                col.source_mappings.append(
                    DHubSourceMapping(
                        source_table_identifier=src_table,
                        source_column_name=src_col,
                    )
                )
            self._model.hubs[hub.base] = dhub
            hub_canon[hub.base.lower()] = hub.base
        return hub_canon

    # -- assembly: links ------------------------------------------------- #

    def _assemble_links(
        self, links: dict[str, _LinkWork], hub_canon: dict[str, str]
    ) -> tuple[dict[str, DLink], dict[str, str]]:
        link_by_iris: dict[str, DLink] = {}
        link_canon: dict[str, str] = {}
        for link in links.values():
            dlink = DLink(physical_name=link.base, link_type="standard")

            # Hub references, ordered by first appearance; resolved to the
            # hub's actual physical name (case-insensitive).
            ref_index_by_hub: dict[str, int] = {}
            for i, hub_iris in enumerate(link.hub_ref_iris):
                stripped = _strip_prefix(hub_iris)[0]
                hub_base = hub_canon.get(stripped.lower(), stripped)
                dlink.hub_references.append(
                    DLinkHubReference(
                        hub_physical_name=hub_base, sort_order=i + 1
                    )
                )
                ref_index_by_hub[hub_iris.lower()] = i

            # Source tables mapping onto the link.
            link_source_tables = {
                src_table.lower()
                for src_table, _src_col, _tgt_col in self._target_mappings(
                    link.iris_name
                )
            }
            # BKCC mappings are dropped from the index, so scan raw link-target
            # rows too to capture the source of BKCC-only links.
            for row in self._mappings_rows:
                v = row.values
                if (clean(v.get("target table")) or "").lower() == link.iris_name.lower():
                    src = clean(v.get("source table"))
                    if src:
                        link_source_tables.add(src.lower())

            # Hub-source mappings: hub-target rows whose source belongs to this
            # link.
            seen_hsm: set[tuple[int, str, str, str]] = set()
            for hub_iris, ref_idx in ref_index_by_hub.items():
                for src_table, src_col, tgt_col in self._target_mappings(hub_iris):
                    if src_table.lower() not in link_source_tables:
                        continue
                    key = (ref_idx, tgt_col.lower(), src_table.lower(), src_col.lower())
                    if key in seen_hsm:
                        continue
                    seen_hsm.add(key)
                    dlink.hub_source_mappings.append(
                        DLinkHubSourceMapping(
                            link_hub_ref_index=ref_idx,
                            hub_column_name=tgt_col,
                            source_table_identifier=src_table,
                            source_column_name=src_col,
                        )
                    )

            # Link-level additional columns.
            for i, col_name in enumerate(link.additional_columns):
                dcol = DLinkColumn(
                    name=col_name, column_type="additional_column", sort_order=i + 1
                )
                for src_table, src_col, tgt_col in self._target_mappings(
                    link.iris_name
                ):
                    if tgt_col.lower() == col_name.lower():
                        dcol.source_mappings.append(
                            DLinkSourceMapping(
                                source_table_identifier=src_table,
                                source_column_name=src_col,
                            )
                        )
                dlink.columns.append(dcol)

            self._model.links[link.base] = dlink
            link_by_iris[link.iris_name.lower()] = dlink
            link_canon[link.base.lower()] = link.base
        return link_by_iris, link_canon

    # -- assembly: satellites ------------------------------------------- #

    def _assemble_satellites(
        self,
        sats: dict[str, _SatWork],
        hub_canon: dict[str, str],
        link_canon: dict[str, str],
        link_by_iris: dict[str, DLink],
    ) -> None:
        used_names: set[str] = set()
        for sat in sats.values():
            stripped, prefix_kind, _ = _strip_prefix(sat.parent_iris)
            parent_lower = stripped.lower()
            # Resolve the parent against the built hubs/links case-insensitively;
            # fall back to the prefix-derived kind for parents with no row.
            if parent_lower in link_canon:
                parent_base, parent_kind = link_canon[parent_lower], "link"
            elif parent_lower in hub_canon:
                parent_base, parent_kind = hub_canon[parent_lower], "hub"
            else:
                parent_base = stripped
                parent_kind = "link" if prefix_kind == "link" else "hub"

            # A satellite on a link whose base matches the link is the exporter's
            # synthesized non-historized payload satellite: fold it into the link.
            if (
                parent_kind == "link"
                and sat.base.lower() == parent_base.lower()
                and sat.parent_iris.lower() in link_by_iris
            ):
                self._fold_payload_satellite(sat, link_by_iris[sat.parent_iris.lower()])
                continue

            self._build_standalone_satellite(sat, parent_base, parent_kind, used_names)

    def _fold_payload_satellite(self, sat: _SatWork, dlink: DLink) -> None:
        dlink.link_type = "non_historized"
        col_source = self._satellite_column_sources(sat)
        start = len(dlink.columns)
        for offset, tgt_col in enumerate(sat.target_columns):
            dcol = DLinkColumn(
                name=tgt_col, column_type="payload", sort_order=start + offset + 1
            )
            src = col_source.get(tgt_col.lower())
            if src is not None:
                dcol.source_mappings.append(
                    DLinkSourceMapping(
                        source_table_identifier=src[0], source_column_name=src[1]
                    )
                )
            dlink.columns.append(dcol)

    def _build_standalone_satellite(
        self,
        sat: _SatWork,
        parent_base: str,
        parent_kind: str | None,
        used_names: set[str],
    ) -> None:
        name = self._unique_sat_name(sat.base, sat.is_multi_active, used_names)
        dsat = DSatellite(
            physical_name=name,
            satellite_type="multi_active" if sat.is_multi_active else "standard",
            parent_entity_name=parent_base,
            parent_entity_type="link" if parent_kind == "link" else "hub",
        )
        col_source = self._satellite_column_sources(sat)
        source_table_id = ""
        for i, tgt_col in enumerate(sat.target_columns):
            src = col_source.get(tgt_col.lower())
            source_col = src[1] if src is not None else tgt_col
            if src is not None and not source_table_id:
                source_table_id = src[0]
            dsat.columns.append(
                DSatelliteColumn(
                    source_column_name=source_col,
                    target_column_name=tgt_col,
                    sort_order=i + 1,
                )
            )
        dsat.source_table_identifier = source_table_id
        self._model.satellites[name] = dsat

    @staticmethod
    def _unique_sat_name(base: str, multi_active: bool, used_names: set[str]) -> str:
        """Return a satellite physical name unique (case-insensitively) within
        ``used_names``, appending ``_ma``/``_s`` (then a counter) on collision."""
        if base.lower() not in used_names:
            used_names.add(base.lower())
            return base
        marker = "_ma" if multi_active else "_s"
        candidate = f"{base}{marker}"
        n = 2
        while candidate.lower() in used_names:
            candidate = f"{base}{marker}_{n}"
            n += 1
        used_names.add(candidate.lower())
        return candidate

    def _satellite_column_sources(
        self, sat: _SatWork
    ) -> dict[str, tuple[str, str]]:
        """Map a satellite's lower-cased target column -> (source_table,
        source_column) from the Mappings file."""
        sources: dict[str, tuple[str, str]] = {}
        for src_table, src_col, tgt_col in self._target_mappings(sat.iris_name):
            sources.setdefault(tgt_col.lower(), (src_table, src_col))
        return sources
