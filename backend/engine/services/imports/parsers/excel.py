"""Excel parser: openpyxl workbook → IRDocument."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from engine.services.imports.errors import Code, PipelineAbort, make_issue
from engine.services.imports.ir import IRDocument, IRRow, IRSheet
from engine.services.imports.parsers.base import clean
from engine.services.imports.types import IssueLocation


def parse_excel(path: Path) -> IRDocument:
    """Read an Excel workbook into an IRDocument.

    Raises PipelineAbort with a single `source.*` Issue if the workbook
    cannot be read at all.
    """
    if not path.exists():
        raise PipelineAbort(
            make_issue(
                severity="error",
                code=Code.SOURCE_UNREADABLE,
                stage="parse",
                message=f"File not found: {path}",
                location=IssueLocation(file=str(path)),
            )
        )

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except (InvalidFileException, OSError, KeyError) as exc:
        raise PipelineAbort(
            make_issue(
                severity="error",
                code=Code.SOURCE_UNREADABLE,
                stage="parse",
                message=f"Could not open Excel file: {exc}",
                location=IssueLocation(file=str(path)),
                suggestion="Confirm the file is a valid .xlsx workbook and not corrupted.",
            )
        ) from exc

    doc = IRDocument(source_name=path.name)
    try:
        if not wb.sheetnames:
            raise PipelineAbort(
                make_issue(
                    severity="error",
                    code=Code.SOURCE_EMPTY,
                    stage="parse",
                    message="The workbook contains no sheets.",
                    location=IssueLocation(file=str(path)),
                )
            )

        for sheet_name in wb.sheetnames:
            sheet = _read_sheet(wb[sheet_name], sheet_name)
            if sheet is not None:
                doc.sheets[sheet_name] = sheet
    finally:
        wb.close()

    return doc


def _read_sheet(ws, sheet_name: str) -> IRSheet | None:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_raw = next(rows_iter)
    except StopIteration:
        return None  # empty sheet — silently skip

    headers = [str(c).strip().lower() for c in header_raw if c is not None]
    if not headers:
        return None

    num_cols = len(headers)
    sheet = IRSheet(name=sheet_name, headers=headers)

    # Header is row 1; first data row is row 2.
    for offset, raw_row in enumerate(rows_iter, start=2):
        trimmed = list(raw_row[:num_cols])
        # Skip fully empty rows so users get clean row numbers in diagnostics.
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in trimmed):
            continue

        values = {}
        for i, col in enumerate(headers):
            cell = trimmed[i] if i < len(trimmed) else None
            values[col] = clean(cell)
        sheet.rows.append(IRRow(row_number=offset, values=values))

    return sheet
