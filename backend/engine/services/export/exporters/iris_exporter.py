"""
IRiS exporter for Data Vault export.

Serializes a `ProjectExport` to the three Excel files that IRiS ingests:

  - Source_<project>.xlsx: source table / column metadata
  - DataVault_<project>.xlsx: Hub / Link / Satellite definitions
  - Mappings_<project>.xlsx: source to data-vault column mappings
"""

from __future__ import annotations

import re
import shutil
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import TYPE_CHECKING
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook

if TYPE_CHECKING:
    from engine.services.export.models import (
        HubDefinition,
        LinkDefinition,
        ProjectExport,
        SatelliteColumnDef,
        SatelliteDefinition,
    )


# IRiS enum values
IRIS_TABLE_TYPE_HUB = "Hub"
IRIS_TABLE_TYPE_LINK = "Link"
IRIS_TABLE_TYPE_SATELLITE = "Satellite"

IRIS_SUBTYPE_MULTI_ACTIVE = "Multi active"
IRIS_SUBTYPE_DEPENDENT_CHILD = "Dependent child"

IRIS_COL_BUSINESS_KEY = "Business key"
IRIS_COL_BKCC = "BKCC"
IRIS_COL_LINK_BUSINESS_KEY = "Link business key"
IRIS_COL_LINK_BKCC = "Link BKCC"
IRIS_COL_CHANGING_ATTRIBUTE = "Changing attribute"
IRIS_COL_DEPENDENT_CHILD_KEY = "Dependent child key"
IRIS_COL_SOURCE_EXTRACT_DATE = "Source extract date"

_ENUM_SHEET_ROWS: list[list[str | None]] = [
    ["Table Types", None, "Column Types", None, "Subtypes"],
    [IRIS_TABLE_TYPE_HUB, None, IRIS_COL_BUSINESS_KEY, None, IRIS_SUBTYPE_MULTI_ACTIVE],
    [
        IRIS_TABLE_TYPE_LINK,
        None,
        IRIS_COL_LINK_BUSINESS_KEY,
        None,
        IRIS_SUBTYPE_DEPENDENT_CHILD,
    ],
    [IRIS_TABLE_TYPE_SATELLITE, None, IRIS_COL_CHANGING_ATTRIBUTE, None, None],
    [None, None, IRIS_COL_DEPENDENT_CHILD_KEY, None, None],
    [None, None, IRIS_COL_BKCC, None, None],
    [None, None, IRIS_COL_LINK_BKCC, None, None],
    [None, None, IRIS_COL_SOURCE_EXTRACT_DATE, None, None],
]

# Satellite types that map to an IRiS row. IRiS doesn't model a
# non-historized satellite separately, so we export those as standard
_SATELLITE_TYPE_TO_SUBTYPE: dict[str, str | None] = {
    "standard": None,
    "non_historized": None,
    "multi_active": IRIS_SUBTYPE_MULTI_ACTIVE,
}

# Satellite types skipped on export
_SATELLITE_TYPES_SKIPPED: set[str] = {
    "effectivity",
    "record_tracking",
}

# Hub types skipped on export
_HUB_TYPES_SKIPPED: set[str] = {"reference"}

# Link types skipped on export
_LINK_TYPES_SKIPPED: set[str] = set()

# Substring-match blacklist applied to TurboVault entity names.
# Names containing these tokens are dropped from the IRiS export
_NAME_BLACKLIST_SUBSTRINGS: tuple[str, ...] = ("duplicate", "test")


def _name_blacklisted(name: str) -> bool:
    lower = name.lower()
    return any(token in lower for token in _NAME_BLACKLIST_SUBSTRINGS)


@dataclass
class IrisExportResult:
    """Outcome of an IRiS export run."""

    success: bool = True
    files: list[Path] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    skipped: list[str] = field(default_factory=list)

    def add_error(self, message: str) -> None:
        self.errors.append(message)
        self.success = False

    def add_warning(self, message: str) -> None:
        self.warnings.append(message)

    def add_skipped(self, message: str) -> None:
        self.skipped.append(message)


class IrisExporter:
    """
    Export a `ProjectExport` to the three IRiS Excel templates.

    Usage:
        exporter = IrisExporter()
        result = exporter.export(project_export, output_dir=Path("./out/iris"))
    """

    def _collect_self_ref_links(
        self,
        project_export: "ProjectExport",
        result: IrisExportResult,
    ) -> set[str]:
        """
        Links referencing the same hub twice
        """
        self_ref: set[str] = set()
        for link in project_export.links:
            if _name_blacklisted(link.link_name):
                continue
            if link.link_type in _LINK_TYPES_SKIPPED:
                continue
            seen_hubs: set[str] = set()
            for ref in link.hub_references:
                if ref.hub_name in seen_hubs:
                    self_ref.add(link.link_name)
                    break
                seen_hubs.add(ref.hub_name)
        for link_name in self_ref:
            result.add_warning(
                f"Self-referencing link {link_name!r} skipped "
            )
        return self_ref

    def _collect_links_with_unresolved_sources(
        self,
        project_export: "ProjectExport",
        result: IrisExportResult,
    ) -> set[str]:
        """
        Links the exporter can't faithfully emit.
        """
        source_cols: dict[str, set[str]] = {}
        for system in project_export.sources:
            for table in system.tables:
                source_cols[table.table_name.lower()] = {
                    c.column_name.lower() for c in table.columns
                }

        # stage_name -> prejoin extraction target aliases (lowercased).
        prejoin_aliases: dict[str, set[str]] = {}
        for stage in project_export.stages:
            aliases = {
                (ext.target_column_alias or ext.source_column_name).lower()
                for prejoin in stage.prejoins
                for ext in prejoin.extraction_columns
                if (ext.target_column_alias or ext.source_column_name)
            }
            if aliases:
                prejoin_aliases[stage.stage_name] = aliases

        skip: dict[str, str] = {}
        for link in project_export.links:
            if link.link_name in self._self_ref_links:
                continue
            for src in link.source_tables:
                cols = source_cols.get(src.source_table.lower(), set())
                stage_aliases = prejoin_aliases.get(src.stage_name, set())
                for m in src.columns:
                    if m.source_column_name.lower() not in cols:
                        skip.setdefault(
                            link.link_name,
                            f"source column {src.source_table}."
                            f"{m.source_column_name} not in source table",
                        )
                    elif m.link_column_name.lower() in stage_aliases:
                        skip.setdefault(
                            link.link_name,
                            "uses a prejoin-extracted column "
                            f"({m.link_column_name})",
                        )

        for link_name, reason in skip.items():
            result.add_warning(
                f"Link {link_name!r} skipped: {reason} "
                "(prejoins are not supported by the IRiS exporter)."
            )
        return set(skip)

    def _collect_emittable_hub_sources(
        self,
        project_export: "ProjectExport",
        result: IrisExportResult,
    ) -> set[tuple[str, str]]:
        """
        `(hub_name, source_table)` pairs to emit, derived without the
        unreliable `is_primary_source` flag. Single-source hubs keep
        their source. Multi-source hubs keep sources whose BK comes from
        a same-named source column (dropping renamed/foreign secondaries);
        if none match, all are kept so the hub isn't left empty.
        """
        def _name_matches(hub: "HubDefinition", src) -> bool:
            if src.column_mappings:
                return any(
                    m.source_column.lower() == m.hub_column.lower()
                    for m in src.column_mappings
                )
            bk_cols = {c.lower() for c in hub.business_key_columns}
            return any(c.lower() in bk_cols for c in src.business_key_columns)

        emittable: set[tuple[str, str]] = set()
        for hub in project_export.hubs:
            if _name_blacklisted(hub.hub_name) or hub.hub_type in _HUB_TYPES_SKIPPED:
                continue
            sources = hub.source_tables
            if len(sources) <= 1:
                for src in sources:
                    emittable.add((hub.hub_name, src.source_table))
                continue
            matching = [s for s in sources if _name_matches(hub, s)]
            kept = matching or list(sources)
            kept_tables = {s.source_table for s in kept}
            for src in sources:
                if src.source_table in kept_tables:
                    emittable.add((hub.hub_name, src.source_table))
                else:
                    result.add_warning(
                        f"Hub {hub.hub_name!r}: secondary source "
                        f"{src.source_table!r} skipped (business key comes "
                        "from a differently-named source column)."
                    )
        return emittable

    def _collect_consumed_hubs(
        self,
        project_export: "ProjectExport",
        result: IrisExportResult,
    ) -> set[str]:
        """
        Hub names referenced by an emitted satellite or link.
        """
        consumed: set[str] = set()

        for sat in project_export.satellites:
            if _name_blacklisted(sat.satellite_name):
                continue
            if sat.satellite_type in _SATELLITE_TYPES_SKIPPED:
                continue
            if sat.satellite_type == "reference":
                continue
            if sat.parent_entity_type == "hub" and sat.parent_entity:
                consumed.add(sat.parent_entity)

        for link in project_export.links:
            if _name_blacklisted(link.link_name):
                continue
            if link.link_type in _LINK_TYPES_SKIPPED:
                continue
            if link.link_name in self._self_ref_links:
                continue
            if link.link_name in self._unresolved_source_links:
                continue
            if len(link.hub_references) < 2:
                continue
            for ref in link.hub_references:
                consumed.add(ref.hub_name)

        for hub in project_export.hubs:
            if _name_blacklisted(hub.hub_name):
                continue
            if hub.hub_type in _HUB_TYPES_SKIPPED:
                continue
            if hub.hub_name not in consumed:
                result.add_warning(
                    f"Hub {hub.hub_name!r}: no satellite or link references "
                    "this hub after IRiS filters."
                )

        return consumed

    def _collect_source_bkccs(self, project_export: "ProjectExport") -> None:
        """
        Populate `_source_bkccs[(source_system, source_table)]` with the
        hub base names for which the source needs a `BKCC_<HUB_BASE>`
        column. A source feeds a hub either directly (HubSourceInfo) or
        indirectly through a Link (LinkSourceInfo + hub_references).
        """
        def _register(system: str, table: str, hub_name: str) -> None:
            iris_hub = _iris_table_name(hub_name, "hub")
            if not iris_hub:
                return
            hub_base = _iris_hub_base_name(iris_hub)
            key = (system, table)
            self._source_bkccs.setdefault(key, set()).add(hub_base)

        for hub in project_export.hubs:
            if _name_blacklisted(hub.hub_name) or hub.hub_type in _HUB_TYPES_SKIPPED:
                continue
            if hub.hub_name not in self._consumed_hubs:
                continue
            for src in hub.source_tables:
                if (hub.hub_name, src.source_table) not in self._emittable_hub_sources:
                    continue
                _register(src.source_system, src.source_table, hub.hub_name)

        for link in project_export.links:
            if _name_blacklisted(link.link_name) or link.link_type in _LINK_TYPES_SKIPPED:
                continue
            if link.link_name in self._self_ref_links:
                continue
            if link.link_name in self._unresolved_source_links:
                continue
            if len(link.hub_references) < 2:
                continue
            for src in link.source_tables:
                for ref in link.hub_references:
                    _register(src.source_system, src.source_table, ref.hub_name)

    def _canonical(self, name: str) -> str:
        if not name:
            return name
        key = name.lower()
        canonical = self._canonical_table_names.setdefault(key, name)
        return canonical

    def export(
        self,
        project_export: ProjectExport,
        output_dir: Path,
        *,
        project_name: str | None = None,
    ) -> IrisExportResult:
        """
        Emit Source_<name>.xlsx, DataVault_<name>.xlsx, Mappings_<name>.xlsx
        into `output_dir`. `project_name` overrides the filename suffix.
        """
        result = IrisExportResult()
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        self._canonical_table_names: dict[str, str] = {}
        self._canonical_source_columns: dict[tuple[str, str], tuple[str, str]] = {}

        # Pre-compute the entity filters the writers consult. Order
        # matters: later sets reference earlier ones.
        self._self_ref_links: set[str] = self._collect_self_ref_links(
            project_export, result
        )
        self._unresolved_source_links: set[str] = (
            self._collect_links_with_unresolved_sources(project_export, result)
        )

        # Non-historized links emit a synthesized payload sat; track them
        # so an explicit NH sat on the same link isn't emitted twice.
        self._nh_link_parents_with_synth_sat: set[str] = {
            link.link_name
            for link in project_export.links
            if link.link_type == "non_historized"
            and link.payload_columns
            and link.link_name not in self._self_ref_links
            and link.link_name not in self._unresolved_source_links
        }

        self._emittable_hub_sources: set[tuple[str, str]] = (
            self._collect_emittable_hub_sources(project_export, result)
        )

        self._consumed_hubs: set[str] = self._collect_consumed_hubs(
            project_export, result
        )

        # Per IRiS, every BK needs a paired BKCC; this maps each source
        # to the hubs it must carry a BKCC_<HUB_BASE> column for.
        self._source_bkccs: dict[tuple[str, str], set[str]] = {}
        self._collect_source_bkccs(project_export)

        name_suffix = _sanitize_for_filename(
            project_name or project_export.project_name
        )

        try:
            source_path = output_dir / f"Source_{name_suffix}.xlsx"
            self._write_source_file(project_export, source_path, result)
            result.files.append(source_path)

            mappings_path = output_dir / f"Mappings_{name_suffix}.xlsx"
            covered_dv_cols = self._write_mappings_file(
                project_export, mappings_path, result
            )
            result.files.append(mappings_path)

            datavault_path = output_dir / f"DataVault_{name_suffix}.xlsx"
            self._write_datavault_file(
                project_export, datavault_path, result, covered_dv_cols
            )
            result.files.append(datavault_path)

            self._record_unsupported(project_export, result)
        except Exception as exc:
            result.add_error(f"IRiS export failed: {exc}")

        return result

    # ------------------------------------------------------------------ #
    # File 1: Source metadata
    # ------------------------------------------------------------------ #
    def _write_source_file(
        self,
        project_export: ProjectExport,
        path: Path,
        result: IrisExportResult,
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Table Schema", "Table Name", "Column", "Datatype", "Size", "Scale"])

        # IRiS rejects duplicate column names within a table; dedupe
        # case-insensitively.
        seen: set[tuple[str, str]] = set()
        for system in project_export.sources:
            schema = _sanitize_iris_name(
                system.schema_name or system.name, _IRIS_TABLE_NAME_MAX
            )
            for table in system.tables:
                table_name = _sanitize_iris_name(
                    table.table_name, _IRIS_TABLE_NAME_MAX
                )
                if not table_name:
                    result.add_warning(
                        f"Source table {table.table_name!r}: name degenerate "
                        "after IRiS sanitisation, skipped"
                    )
                    continue
                for column in table.columns:
                    col_name = _sanitize_iris_name(
                        column.column_name, _IRIS_COLUMN_NAME_MAX
                    )
                    if not col_name:
                        continue
                    key = (table_name.lower(), col_name.lower())
                    if key in seen:
                        result.add_warning(
                            f"Source {table_name!r} column "
                            f"{column.column_name!r}: duplicate name. "
                            "Skipped for IRiS export."
                        )
                        continue
                    seen.add(key)

                    self._canonical_source_columns[key] = (
                        table_name,
                        col_name,
                    )
                    datatype, size, scale = _split_datatype(column.datatype)
                    ws.append(
                        [
                            schema,
                            table_name,
                            col_name,
                            datatype,
                            size,
                            scale,
                        ]
                    )

                # Synthesize BKCC_<HUB_BASE> columns for every hub this
                # source feeds (directly or via a link). Mappings + DV
                # writers emit matching column rows that reference these.
                for hub_base in sorted(
                    self._source_bkccs.get((system.name, table.table_name), set())
                ):
                    bkcc_col = _sanitize_iris_name(
                        _link_bkcc_column_name(hub_base), _IRIS_COLUMN_NAME_MAX
                    )
                    if not bkcc_col:
                        continue
                    key = (table_name.lower(), bkcc_col.lower())
                    if key in seen:
                        continue
                    seen.add(key)
                    self._canonical_source_columns[key] = (table_name, bkcc_col)
                    ws.append(
                        [
                            schema,
                            table_name,
                            bkcc_col,
                            "varchar",
                            _BKCC_SOURCE_SIZE,
                            None,
                        ]
                    )

        wb.save(path)
        # IRiS's xlsx parser only reads shared strings; openpyxl writes
        # inline strings by default, so convert. Verified: IRiS rejects
        # the inline-string form.
        _convert_inline_to_shared_strings(path)

    # ------------------------------------------------------------------ #
    # File 2: Data Vault metadata
    # ------------------------------------------------------------------ #
    def _write_datavault_file(
        self,
        project_export: ProjectExport,
        path: Path,
        result: IrisExportResult,
        covered_dv_cols: set[tuple[str, str]],
    ) -> None:
        wb = Workbook()
        data_sheet = wb.active
        data_sheet.title = "Sheet1"
        data_sheet.append(
            [
                "Table Type",
                "Subtype",
                "Table Name",
                "Column",
                "Datatype",
                "Size",
                "Scale",
                "Column Types",
                "Parent Table",
                "Relationship",
                "Relationship Name",
            ]
        )

        seen: set[tuple[str, str]] = set()

        for hub in project_export.hubs:
            self._append_hub_rows(data_sheet, hub, result, seen, covered_dv_cols)

        hubs_emitted = {table.lower() for (table, _col) in seen}

        hub_lookup = {hub.hub_name: hub for hub in project_export.hubs}

        for link in project_export.links:
            self._append_link_rows(
                data_sheet,
                link,
                result,
                seen,
                covered_dv_cols,
                hubs_emitted,
                hub_lookup,
            )

        parents_emitted = {table.lower() for (table, _col) in seen}

        for link in project_export.links:
            if link.link_type != "non_historized":
                continue
            if _name_blacklisted(link.link_name):
                continue
            self._append_non_historized_link_payload_rows(
                data_sheet, link, result, seen, covered_dv_cols, parents_emitted
            )

        for sat in project_export.satellites:
            self._append_satellite_rows(
                data_sheet, sat, result, seen, covered_dv_cols, parents_emitted
            )

        enum_sheet = wb.create_sheet("Sheet2")
        for row in _ENUM_SHEET_ROWS:
            enum_sheet.append(row)

        wb.save(path)
        # IRiS's xlsx parser only reads shared strings
        # openpyxl writes inline strings by default
        _convert_inline_to_shared_strings(path)

    def _append_hub_rows(
        self,
        sheet,
        hub: HubDefinition,
        result: IrisExportResult,
        seen: set[tuple[str, str]],
        covered_dv_cols: set[tuple[str, str]],
    ) -> None:
        if _name_blacklisted(hub.hub_name):
            result.add_warning(
                f"Hub {hub.hub_name!r}: IRiS has no input representation for this hub"
            )
            return

        if hub.hub_type in _HUB_TYPES_SKIPPED:
            result.add_warning(
                f"Reference hub {hub.hub_name!r}: reference entities are not "
                "forwarded to IRiS. Model as a regular hub instead"
            )
            return

        if hub.hub_name not in self._consumed_hubs:
            return

        hub_name = self._canonical(_iris_table_name(hub.hub_name, "hub"))
        if not hub_name:
            result.add_warning(
                f"Hub {hub.hub_name!r}: name degenerate after IRiS "
                "sanitisation, skipped"
            )
            return
        bk_columns = hub.business_key_columns or hub.reference_key_columns
        bk_set = set(hub.business_key_columns) | set(hub.reference_key_columns)

        for raw_col in bk_columns:
            col_name = _sanitize_iris_name(raw_col, _IRIS_COLUMN_NAME_MAX)
            if not col_name:
                continue
            if (hub_name.lower(), col_name.lower()) not in covered_dv_cols:
                continue
            if _dedupe_skip(seen, hub_name, col_name, result, "Hub"):
                continue
            datatype, size, scale = _bk_datatype(
                None,
                entity=hub_name,
                column=col_name,
                result=result,
            )
            sheet.append(
                [
                    IRIS_TABLE_TYPE_HUB,
                    None,
                    hub_name,
                    col_name,
                    datatype,
                    size,
                    scale,
                    IRIS_COL_BUSINESS_KEY,
                    None,
                    None,
                    None,
                ]
            )

        # IRiS expects every hub to carry a `BKCC` collision-code column
        # alongside its business key. Emit one synthesized row per hub.
        if (hub_name.lower(), _HUB_BKCC_COLUMN.lower()) in covered_dv_cols:
            if not _dedupe_skip(seen, hub_name, _HUB_BKCC_COLUMN, result, "Hub"):
                sheet.append(
                    [
                        IRIS_TABLE_TYPE_HUB,
                        None,
                        hub_name,
                        _HUB_BKCC_COLUMN,
                        "varchar",
                        _BKCC_DV_SIZE,
                        None,
                        IRIS_COL_BKCC,
                        None,
                        None,
                        None,
                    ]
                )

        # Non-key columns on a hub (typically BKCC / source-extract-date)
        # are inferred by naming convention.
        for raw_col in hub.additional_columns:
            if raw_col in bk_set:
                continue
            column_type = _guess_hub_additional_column_type(raw_col)
            if column_type is None:
                result.add_warning(
                    f"Hub {hub_name!r} column {raw_col!r}: "
                    "could not infer IRiS Column Type, skipping"
                )
                continue
            col_name = _sanitize_iris_name(raw_col, _IRIS_COLUMN_NAME_MAX)
            if not col_name:
                continue
            if (hub_name.lower(), col_name.lower()) not in covered_dv_cols:
                continue
            if _dedupe_skip(seen, hub_name, col_name, result, "Hub"):
                continue
            sheet.append(
                [
                    IRIS_TABLE_TYPE_HUB,
                    None,
                    hub_name,
                    col_name,
                    "varchar",
                    _DEFAULT_VARCHAR_SIZE,
                    None,
                    column_type,
                    None,
                    None,
                    None,
                ]
            )

    def _append_link_rows(
        self,
        sheet,
        link: LinkDefinition,
        result: IrisExportResult,
        seen: set[tuple[str, str]],
        covered_dv_cols: set[tuple[str, str]],
        hubs_emitted: set[str],
        hub_lookup: dict[str, "HubDefinition"],
    ) -> None:

        if _name_blacklisted(link.link_name):
            result.add_warning(
                f"Link {link.link_name!r}: IRiS has no input representation for this link"
            )
            return

        if link.link_type in _LINK_TYPES_SKIPPED:
            result.add_warning(
                f"Link {link.link_name!r} (type={link.link_type}): "
                "IRiS has no input representation for this link type"
            )
            return

        if link.link_name in self._self_ref_links:
            return

        if link.link_name in self._unresolved_source_links:
            return

        link_name = self._canonical(_iris_table_name(link.link_name, "link"))
        if not link_name:
            result.add_warning(
                f"Link {link.link_name!r}: name degenerate after IRiS "
                "sanitisation, skipped"
            )
            return
        if len(link.hub_references) < 2:
            result.add_warning(
                f"Link {link_name!r}: fewer than 2 hub references. "
                "IRiS requires Links to connect at least two hubs."
            )
            return

        for ref in link.hub_references:
            related_hub_name = self._canonical(
                _iris_table_name(ref.hub_name, "hub")
            )
            if not related_hub_name:
                continue
            if related_hub_name.lower() not in hubs_emitted:
                continue
            hub_def = hub_lookup.get(ref.hub_name)
            if hub_def is None:
                result.add_warning(
                    f"Link {link_name!r}: referenced hub {ref.hub_name!r} "
                    "not found in project, skipping that connection"
                )
                continue

            hub_base = _iris_hub_base_name(related_hub_name)
            relationship_name = hub_base

            # Paired LINK BKCC column for this hub reference.
            link_bkcc_col = _link_bkcc_column_name(hub_base)
            if (link_name.lower(), link_bkcc_col.lower()) in covered_dv_cols:
                if not _dedupe_skip(seen, link_name, link_bkcc_col, result, "Link"):
                    sheet.append(
                        [
                            IRIS_TABLE_TYPE_LINK,
                            None,
                            link_name,
                            link_bkcc_col,
                            "varchar",
                            _BKCC_DV_SIZE,
                            None,
                            IRIS_COL_LINK_BKCC,
                            None,
                            related_hub_name,
                            relationship_name,
                        ]
                    )

            for raw_bk in hub_def.business_key_columns or hub_def.reference_key_columns:
                bk_column = _sanitize_iris_name(raw_bk, _IRIS_COLUMN_NAME_MAX)
                if not bk_column:
                    continue
                if (link_name.lower(), bk_column.lower()) not in covered_dv_cols:
                    continue
                if _dedupe_skip(seen, link_name, bk_column, result, "Link"):
                    continue
                sheet.append(
                    [
                        IRIS_TABLE_TYPE_LINK,
                        None,
                        link_name,
                        bk_column,
                        "varchar",
                        _DEFAULT_VARCHAR_SIZE,
                        None,
                        IRIS_COL_LINK_BUSINESS_KEY,
                        None,
                        related_hub_name,
                        relationship_name,
                    ]
                )

        # Link-level additional columns
        for raw_col in link.additional_columns:
            column_type = _guess_link_additional_column_type(raw_col)
            if column_type is None:
                result.add_warning(
                    f"Link {link_name!r} column {raw_col!r}: "
                    "could not infer IRiS Column Type, skipping"
                )
                continue
            if column_type == IRIS_COL_LINK_BKCC:
                continue
            col_name = _sanitize_iris_name(raw_col, _IRIS_COLUMN_NAME_MAX)
            if not col_name:
                continue
            if (link_name.lower(), col_name.lower()) not in covered_dv_cols:
                continue
            if _dedupe_skip(seen, link_name, col_name, result, "Link"):
                continue
            sheet.append(
                [
                    IRIS_TABLE_TYPE_LINK,
                    None,
                    link_name,
                    col_name,
                    "varchar",
                    _DEFAULT_VARCHAR_SIZE,
                    None,
                    column_type,
                    None,
                    None,
                    None,
                ]
            )

    def _append_non_historized_link_payload_rows(
        self,
        sheet,
        link: LinkDefinition,
        result: IrisExportResult,
        seen: set[tuple[str, str]],
        covered_dv_cols: set[tuple[str, str]],
        parents_emitted: set[str],
    ) -> None:
        """
        Emit a synthesized Satellite carrying a non-historized link's payload.

        IRiS doesn't model "non-historized link" as a distinct concept, so
        the link itself is exported as a standard Link and Satellite with payload
        """
        if not link.payload_columns:
            return

        parent_name = self._canonical(_iris_table_name(link.link_name, "link"))
        if not parent_name or parent_name.lower() not in parents_emitted:
            result.add_warning(
                f"Non-historized link {link.link_name!r}: parent link did "
                "not emit rows, skipping synthesized payload satellite"
            )
            return

        sat_name = self._canonical(_link_payload_satellite_name(parent_name))
        if not sat_name:
            return

        for raw_col in link.payload_columns:
            col_name = _sanitize_iris_name(raw_col, _IRIS_COLUMN_NAME_MAX)
            if not col_name:
                continue
            if (sat_name.lower(), col_name.lower()) not in covered_dv_cols:
                continue
            if _dedupe_skip(seen, sat_name, col_name, result, "Satellite"):
                continue
            sheet.append(
                [
                    IRIS_TABLE_TYPE_SATELLITE,
                    None,
                    sat_name,
                    col_name,
                    "varchar",
                    _DEFAULT_VARCHAR_SIZE,
                    None,
                    IRIS_COL_CHANGING_ATTRIBUTE,
                    parent_name,
                    None,
                    None,
                ]
            )

    def _append_satellite_rows(
        self,
        sheet,
        sat: SatelliteDefinition,
        result: IrisExportResult,
        seen: set[tuple[str, str]],
        covered_dv_cols: set[tuple[str, str]],
        parents_emitted: set[str],
    ) -> None:
        if _name_blacklisted(sat.satellite_name):
            result.add_warning(
                f"Satellite {sat.satellite_name!r}: IRiS has no input representation for this satellite"
            )
            return

        if sat.satellite_type == "reference":
            result.add_warning(
                f"Reference satellite {sat.satellite_name!r}: reference "
                "entities are not forwarded to IRiS. Model as a regular "
                "satellite instead"
            )
            return

        if sat.satellite_type in _SATELLITE_TYPES_SKIPPED:
            result.add_warning(
                f"Satellite {sat.satellite_name!r} (type={sat.satellite_type}): "
                "IRiS auto-generates this type from the parent hub"
            )
            return

        # Non-historized sats whose parent is a non-historized link
        # would duplicate the synthesized payload sat that
        # `_append_non_historized_link_payload_rows` already emits.
        if (
            sat.satellite_type == "non_historized"
            and sat.parent_entity_type == "link"
            and sat.parent_entity in self._nh_link_parents_with_synth_sat
        ):
            result.add_warning(
                f"Non-historized satellite {sat.satellite_name!r}: parent "
                f"non-historized link {sat.parent_entity!r} already emits a "
                "satellite covering the same target. "
                "Skipping to avoid duplicate columns in IRiS."
            )
            return

        if not sat.parent_entity:
            # Parent Table is mandatory per the IRiS spec for Satellite rows.
            result.add_warning(
                f"Satellite {sat.satellite_name!r}: no parent entity. "
                "IRiS requires a Parent Table for every Satellite row."
            )
            return

        sat_name = self._canonical(
            _iris_table_name(sat.satellite_name, "satellite", sat.satellite_type)
        )
        parent_name = self._canonical(
            _iris_table_name(sat.parent_entity, sat.parent_entity_type)
        )
        if not sat_name or not parent_name:
            result.add_warning(
                f"Satellite {sat.satellite_name!r}: name(s) degenerate after "
                "IRiS sanitisation, skipped"
            )
            return

        # If the parent hub/link didn't actually emit any rows above,
        # IRiS would reject this Satellite's Parent Table reference.
        if parent_name.lower() not in parents_emitted:
            result.add_warning(
                f"Satellite {sat.satellite_name!r}: parent "
                f"{parent_name!r} is not in the DV file, skipped"
            )
            return

        subtype = _SATELLITE_TYPE_TO_SUBTYPE.get(sat.satellite_type)
        if sat.satellite_type not in _SATELLITE_TYPE_TO_SUBTYPE:
            result.add_warning(
                f"Satellite {sat_name!r}: unknown type "
                f"{sat.satellite_type!r}, exporting as plain Satellite"
            )

        for col in sat.columns:
            raw_target = col.target_column_name or col.source_column
            target_name = _sanitize_iris_name(raw_target, _IRIS_COLUMN_NAME_MAX)
            if not target_name:
                continue
            if (sat_name.lower(), target_name.lower()) not in covered_dv_cols:
                continue
            if _dedupe_skip(seen, sat_name, target_name, result, "Satellite"):
                continue
            datatype, size, scale = _split_datatype(_satellite_column_datatype(col))
            # IRiS only allows 'Dependent child key' for Subtype
            # 'Dependent child'. Multi active sats (and standard ones)
            # use 'Changing attribute' for every payload column.
            column_type = IRIS_COL_CHANGING_ATTRIBUTE
            sheet.append(
                [
                    IRIS_TABLE_TYPE_SATELLITE,
                    subtype,
                    sat_name,
                    target_name,
                    datatype,
                    size,
                    scale,
                    column_type,
                    parent_name,
                    None,
                    None,
                ]
            )

    # ------------------------------------------------------------------ #
    # File 3: Source-to-DV mappings
    # ------------------------------------------------------------------ #
    def _write_mappings_file(
        self,
        project_export: ProjectExport,
        path: Path,
        result: IrisExportResult,
    ) -> set[tuple[str, str]]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(
            [
                "Source Table",
                "Source Column",
                "Target Table",
                "Target Column",
                "Mapping Set Name",
            ]
        )

        # Dedupe full mapping rows only
        seen_rows: set[tuple[str, str, str, str, str]] = set()
        # Record every (target_table, target_column) we cover, so the
        # DV emission can filter itself to match.
        covered: set[tuple[str, str]] = set()

        def _append(
            source_table: str,
            source_column: str,
            target_table: str,
            target_column: str,
            mapping_set: str,
        ) -> None:
            # Mapping cells must satisfy the same IRiS naming rules as
            # the Source / DV files they reference
            source_table = _sanitize_iris_name(source_table, _IRIS_TABLE_NAME_MAX)
            source_column = _sanitize_iris_name(source_column, _IRIS_COLUMN_NAME_MAX)
            target_table = _sanitize_iris_name(target_table, _IRIS_TABLE_NAME_MAX)
            target_column = _sanitize_iris_name(target_column, _IRIS_COLUMN_NAME_MAX)
            mapping_set = _sanitize_mapping_set(mapping_set)
            # If the Source writer kept a different casing for this
            # (table, column), rewrite to that casing so IRiS auto-link
            # finds an exact match in the Source file.
            canonical = self._canonical_source_columns.get(
                (source_table.lower(), source_column.lower())
            )
            if canonical is not None:
                source_table, source_column = canonical
            if not all(
                [source_table, source_column, target_table, target_column, mapping_set]
            ):
                return

            key = (
                source_table.lower(),
                source_column.lower(),
                target_table.lower(),
                target_column.lower(),
                mapping_set.lower(),
            )
            if key in seen_rows:
                covered.add((target_table.lower(), target_column.lower()))
                return
            covered.add((target_table.lower(), target_column.lower()))
            seen_rows.add(key)
            ws.append(
                [
                    source_table,
                    source_column,
                    target_table,
                    target_column,
                    mapping_set,
                ]
            )

        for hub in project_export.hubs:
            if hub.hub_type in _HUB_TYPES_SKIPPED or _name_blacklisted(hub.hub_name):
                continue
            if hub.hub_name not in self._consumed_hubs:
                continue
            hub_name = self._canonical(_iris_table_name(hub.hub_name, "hub"))
            hub_base = _iris_hub_base_name(hub_name)
            hub_bkcc_src_col = _link_bkcc_column_name(hub_base)
            for source in hub.source_tables:
                if (hub.hub_name, source.source_table) not in self._emittable_hub_sources:
                    continue
                mset = _mapping_set_name(
                    source.source_system, source.source_table, hub_name
                )
                _append(
                    source.source_table,
                    hub_bkcc_src_col,
                    hub_name,
                    _HUB_BKCC_COLUMN,
                    mset,
                )
                if source.column_mappings:
                    for m in source.column_mappings:
                        _append(
                            source.source_table,
                            m.source_column,
                            hub_name,
                            m.hub_column,
                            mset,
                        )
                else:
                    for hub_bk, src_bk in zip(
                        hub.business_key_columns,
                        source.business_key_columns,
                        strict=False,
                    ):
                        _append(
                            source.source_table,
                            src_bk,
                            hub_name,
                            hub_bk,
                            mset,
                        )

        hub_lookup = {hub.hub_name: hub for hub in project_export.hubs}

        for link in project_export.links:
            if link.link_type in _LINK_TYPES_SKIPPED or _name_blacklisted(
                link.link_name
            ):
                continue
            if link.link_name in self._self_ref_links:
                continue
            if link.link_name in self._unresolved_source_links:
                continue
            if len(link.hub_references) < 2:
                continue
            link_name = self._canonical(_iris_table_name(link.link_name, "link"))

            payload_sat_name = (
                self._canonical(_link_payload_satellite_name(link_name))
                if link.link_type == "non_historized"
                else None
            )
            for source in link.source_tables:
                link_mset = _mapping_set_name(
                    source.source_system, source.source_table, link_name
                )
                payload_mset = (
                    _mapping_set_name(
                        source.source_system,
                        source.source_table,
                        payload_sat_name,
                    )
                    if payload_sat_name is not None
                    else None
                )

                # Synthesize secondary hub source mappings
                for ref in link.hub_references:
                    ref_iris_hub = self._canonical(
                        _iris_table_name(ref.hub_name, "hub")
                    )
                    if not ref_iris_hub:
                        continue
                    hub_def = hub_lookup.get(ref.hub_name)
                    if hub_def is None:
                        continue
                    ref_hub_base = _iris_hub_base_name(ref_iris_hub)
                    bkcc_col = _link_bkcc_column_name(ref_hub_base)
                    hub_mset = _mapping_set_name(
                        source.source_system,
                        source.source_table,
                        ref_iris_hub,
                    )
                    hub_bks = list(
                        hub_def.business_key_columns
                        or hub_def.reference_key_columns
                    )
                    hub_bk_lookup = {bk.lower(): bk for bk in hub_bks}
                    for m in source.columns:
                        if m.link_column_type != "business_key":
                            continue
                        canonical_bk = hub_bk_lookup.get(m.link_column_name.lower())
                        if canonical_bk is None:
                            continue
                        _append(
                            source.source_table,
                            m.source_column_name,
                            ref_iris_hub,
                            canonical_bk,
                            hub_mset,
                        )
                    _append(
                        source.source_table,
                        bkcc_col,
                        ref_iris_hub,
                        _HUB_BKCC_COLUMN,
                        hub_mset,
                    )

                # Per-hub Link BKCC mapping. The link's own BKCC column
                # for each related hub draws from the same source BKCC
                # column the hub uses.
                for ref in link.hub_references:
                    ref_iris_hub = _iris_table_name(ref.hub_name, "hub")
                    if not ref_iris_hub:
                        continue
                    ref_hub_base = _iris_hub_base_name(ref_iris_hub)
                    bkcc_col = _link_bkcc_column_name(ref_hub_base)
                    _append(
                        source.source_table,
                        bkcc_col,
                        link_name,
                        bkcc_col,
                        link_mset,
                    )
                for m in source.columns:
                    if m.link_column_type == "business_key" or (
                        _guess_link_additional_column_type(m.link_column_name)
                    ):
                        _append(
                            source.source_table,
                            m.source_column_name,
                            link_name,
                            m.link_column_name,
                            link_mset,
                        )
                    elif (
                        m.link_column_type == "payload"
                        and payload_sat_name is not None
                        and payload_mset is not None
                    ):
                        _append(
                            source.source_table,
                            m.source_column_name,
                            payload_sat_name,
                            m.link_column_name,
                            payload_mset,
                        )

        for sat in project_export.satellites:
            if sat.satellite_type in _SATELLITE_TYPES_SKIPPED or _name_blacklisted(
                sat.satellite_name
            ):
                continue
            # Reference satellites are warned-and-skipped by the DV writer
            if sat.satellite_type == "reference":
                continue
            # Non-historized sats whose parent NH link emits a synth
            # payload sat are skipped by the DV writer
            if (
                sat.satellite_type == "non_historized"
                and sat.parent_entity_type == "link"
                and sat.parent_entity in self._nh_link_parents_with_synth_sat
            ):
                continue
            sat_name = self._canonical(
                _iris_table_name(
                    sat.satellite_name, "satellite", sat.satellite_type
                )
            )
            mset = _mapping_set_name(
                sat.source_system, sat.source_table, sat_name
            )
            for col in sat.columns:
                target_name = col.target_column_name or col.source_column
                _append(
                    sat.source_table,
                    col.source_column,
                    sat_name,
                    target_name,
                    mset,
                )

        wb.save(path)
        # IRiS's xlsx parser only reads shared strings; openpyxl writes
        # inline strings by default, so convert
        _convert_inline_to_shared_strings(path)
        return covered

    # ------------------------------------------------------------------ #
    # Unsupported-feature reporting
    # ------------------------------------------------------------------ #
    def _record_unsupported(
        self,
        project_export: ProjectExport,
        result: IrisExportResult,
    ) -> None:
        for pit in project_export.pits:
            result.add_warning(
                f"PIT {pit.pit_name!r}: PIT entities are not forwarded to IRiS"
            )
        for ref in project_export.reference_tables:
            result.add_warning(
                f"Reference table {ref.table_name!r}: reference entities are "
                "not forwarded to IRiS. Model as a regular hub + satellite "
                "instead"
            )


# ====================================================================== #
# Helper functions
# ====================================================================== #


# IRiS spec: Table Name allows alphanumeric + '-' + '_' (max 122 chars,
# no spaces). Column allows the same set but up to 128.
_NAME_INVALID_CHARS_RE = re.compile(r"[^A-Za-z0-9_\-]")
_IRIS_TABLE_NAME_MAX = 122
_IRIS_COLUMN_NAME_MAX = 128
_DEFAULT_VARCHAR_SIZE = 4000
_BKCC_SOURCE_SIZE = 20
_BKCC_DV_SIZE = 40

_HUB_BKCC_COLUMN = "BKCC"


def _iris_hub_base_name(hub_iris_name: str) -> str:
    """Strip the `h_` prefix from a hub's IRiS name to get the bare base."""
    if hub_iris_name.lower().startswith("h_"):
        return hub_iris_name[2:]
    return hub_iris_name


def _link_bkcc_column_name(hub_base: str) -> str:
    """Link-side BKCC column name for a reference to a hub"""
    return f"BKCC_{hub_base}"


def _sanitize_iris_name(name: str, max_len: int) -> str:
    """
    Apply IRiS character/length rules
    """
    if not name:
        return ""
    collapsed = name.replace(" ", "_")
    cleaned = _NAME_INVALID_CHARS_RE.sub("", collapsed)
    return cleaned[:max_len]


def _iris_table_name(
    name: str,
    entity_type: str,
    satellite_type: str | None = None,
) -> str:
    """
    Normalize a Hub/Link/Satellite name to the IRiS prefix convention:
      - h_<...>       Hub
      - l_<...>       Link
      - s_<...>       Standard / reference satellite
      - s_ma_<...>    Multi active satellite

    TurboVault's inverse suffixes (`_H`/`_L`/`_NL`/`_S`/`_MS`/`_NHS`) are
    stripped before prefixing to avoid names like `h_CUSTOMER_H`. The
    `_p`/`_n` modifier is kept so each sat maps to a distinct IRiS sat.
    Returns "" if no legal characters survive.
    """
    if entity_type == "hub":
        prefix = "h_"
        suffixes: tuple[str, ...] = ("_h",)
    elif entity_type == "link":
        prefix = "l_"
        # `_nl` (non-historized link) listed before `_l` so it wins.
        suffixes = ("_nhl", "_nl", "_l")
    elif entity_type == "satellite":
        if satellite_type == "multi_active":
            prefix = "s_ma_"
            suffixes = ("_mas", "_ms")
        else:
            prefix = "s_"
            suffixes = ("_nhs", "_ns", "_s")
    else:
        return _sanitize_iris_name(name, _IRIS_TABLE_NAME_MAX)

    if name.lower().startswith(prefix):
        return _sanitize_iris_name(name, _IRIS_TABLE_NAME_MAX)

    base = name
    lower = name.lower()
    for suffix in suffixes:
        if lower.endswith(suffix) and len(base) > len(suffix):
            base = base[: -len(suffix)]
            break

    return _sanitize_iris_name(f"{prefix}{base}", _IRIS_TABLE_NAME_MAX)


def _link_payload_satellite_name(link_iris_name: str) -> str:
    """
    Satellite name carrying a non-historized link's payload: swap the
    link's `l_` prefix for `s_` (e.g. `l_PART_SUPPLIER` to `s_PART_SUPPLIER`).
    """
    if link_iris_name.lower().startswith("l_"):
        base = link_iris_name[2:]
    else:
        base = link_iris_name
    return _sanitize_iris_name(f"s_{base}", _IRIS_TABLE_NAME_MAX)


def _sanitize_for_filename(name: str) -> str:
    """Strip characters that confuse downstream consumers / filesystems."""
    return "".join(c if (c.isalnum() or c in ("-", "_")) else "_" for c in name)


def _sanitize_mapping_set(name: str) -> str:
    """Mapping Set Name in IRiS allows alphanumeric + hyphen + underscore."""
    return "".join(c if (c.isalnum() or c in ("-", "_")) else "_" for c in name)


def _mapping_set_name(
    source_system: str, source_table: str, target_table: str
) -> str:
    """
    Mapping Set Name keyed by source table
    """
    del target_table
    sys_part = _sanitize_mapping_set(source_system or "")
    tbl_part = _sanitize_mapping_set(source_table or "")
    if tbl_part and sys_part:
        return f"{tbl_part}_{sys_part}"
    return tbl_part or sys_part


def _split_datatype(raw: str | None) -> tuple[str, int | None, int | None]:
    """
    Split a TurboVault datatype string (e.g. 'decimal(18,4)') into IRiS's
    separate (Datatype, Size, Scale) cells.
    """
    if not raw:
        return "varchar", _DEFAULT_VARCHAR_SIZE, None

    text = raw.strip()
    if "(" in text and text.endswith(")"):
        base, _, rest = text.partition("(")
        inner = rest[:-1]
        parts = [p.strip() for p in inner.split(",")]
        size: int | None = None
        scale: int | None = None
        if parts:
            size = _parse_size(parts[0])
        if len(parts) > 1:
            try:
                scale = int(parts[1])
            except ValueError:
                scale = None
        return base.strip(), size, scale

    return text, None, None


def _parse_size(value: str) -> int | None:
    """
    Parse a SQL-style size component into an IRiS-compatible integer.

    IRiS only accepts numeric sizes in the Size cell; the SQL Server
    'max' keyword is mapped to _DEFAULT_VARCHAR_SIZE.
    """
    if not value:
        return None
    if value.lower() == "max":
        return _DEFAULT_VARCHAR_SIZE
    try:
        return int(value)
    except ValueError:
        return None


def _bk_datatype(
    raw: str | None,
    *,
    entity: str,
    column: str,
    result: IrisExportResult,
) -> tuple[str, int | None, int | None]:
    """
    Resolve the datatype triple for a (Link) Business key column.

    IRiS mandates VARCHAR for BK columns to enable Null BK substitution
    and Ghost record processing. We coerce non-varchar inputs and warn.
    """
    base, size, _ = _split_datatype(raw)
    if base.lower().startswith("varchar"):
        return "varchar", size if size is not None else _DEFAULT_VARCHAR_SIZE, None
    if raw is not None:
        result.add_warning(
            f"{entity}:{column}: coerced datatype {raw!r} to "
            f"varchar({_DEFAULT_VARCHAR_SIZE}). "
            "IRiS requires VARCHAR for business keys."
        )
    return "varchar", _DEFAULT_VARCHAR_SIZE, None


def _satellite_column_datatype(satellite_column: SatelliteColumnDef) -> str:
    """
    Best-effort datatype extraction for a SatelliteColumnDef.

    The export model doesn't carry a per-column datatype on satellites,
    only on stage / source columns. Default to varchar(max); IRiS will
    not reject this for changing attributes.
    """
    return "varchar(max)"


def _guess_hub_additional_column_type(column_name: str) -> str | None:
    """
    Infer the IRiS `Column Types` value for a non-BK hub column.

    Heuristic based on naming conventions used in datavault4dbt pipelines.
    Returns None when we can't make a confident guess.
    """
    lower = column_name.lower()
    if "bkcc" in lower or "collision" in lower:
        return IRIS_COL_BKCC
    if "extract" in lower or "load_date" in lower or "ldts" in lower:
        return IRIS_COL_SOURCE_EXTRACT_DATE
    return None


def _dedupe_skip(
    seen: set[tuple[str, str]],
    table_name: str,
    column_name: str,
    result: IrisExportResult,
    entity_kind: str,
) -> bool:
    """
    Track (table, column) pairs case-insensitively across the DataVault
    sheet. Returns True (and warns) on a repeat so the caller skips it.
    """
    key = (table_name.lower(), column_name.lower())
    if key in seen:
        result.add_warning(
            f"{entity_kind} {table_name!r} column {column_name!r}: "
            "duplicate name. Skipped for IRiS export."
        )
        return True
    seen.add(key)
    return False


def _guess_link_additional_column_type(column_name: str) -> str | None:
    """Same idea as the hub helper, scoped to link Column Types."""
    lower = column_name.lower()
    if "bkcc" in lower or "collision" in lower:
        return IRIS_COL_LINK_BKCC
    if "extract" in lower or "load_date" in lower or "ldts" in lower:
        return IRIS_COL_SOURCE_EXTRACT_DATE
    return None


# ====================================================================== #
# Shared-strings post-processor
# ====================================================================== #
#
# openpyxl writes cell text as inline strings, but IRiS's xlsx parser
# only reads the shared-strings form. We rewrite the file afterward,
# handling just the inline-string shape openpyxl produces.

_INLINE_CELL_RE = re.compile(
    r'<c([^>]*)\st="inlineStr"([^>]*)>\s*<is>\s*<t(?:\s[^>]*)?>([^<]*)</t>\s*</is>\s*</c>'
)


def _convert_inline_to_shared_strings(xlsx_path: Path) -> None:
    """
    Rewrite an openpyxl-produced xlsx so cell text is stored as shared
    strings instead of inline strings. Mutates the file in place.
    """
    shared_table: dict[str, int] = {}
    ordered_strings: list[str] = []

    def _intern(text: str) -> int:
        if text not in shared_table:
            shared_table[text] = len(ordered_strings)
            ordered_strings.append(text)
        return shared_table[text]

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        members = {name: zf.read(name) for name in zf.namelist()}

    # 1. Rewrite each sheet, replacing inline strings with shared refs.
    sheet_names = [n for n in members if n.startswith("xl/worksheets/sheet")]
    for name in sheet_names:
        xml = members[name].decode("utf-8")

        def _replace(match: re.Match[str]) -> str:
            pre_attrs, post_attrs, raw_text = match.groups()
            # The captured text was XML-escaped when openpyxl wrote it;
            # we re-escape what we put into sharedStrings.xml so the
            # round-trip is byte-stable.
            idx = _intern(raw_text)
            return f'<c{pre_attrs}{post_attrs} t="s"><v>{idx}</v></c>'

        members[name] = _INLINE_CELL_RE.sub(_replace, xml).encode("utf-8")

    # 2. Build xl/sharedStrings.xml.
    if ordered_strings:
        si_blocks = "".join(
            f"<si><t xml:space=\"preserve\">{_text_xml(s)}</t></si>"
            for s in ordered_strings
        )
        members["xl/sharedStrings.xml"] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            f'count="{len(ordered_strings)}" uniqueCount="{len(ordered_strings)}">'
            f"{si_blocks}</sst>"
        ).encode("utf-8")

        # 3. Declare the part in [Content_Types].xml.
        ct_path = "[Content_Types].xml"
        ct_xml = members[ct_path].decode("utf-8")
        if "sharedStrings.xml" not in ct_xml:
            override = (
                '<Override PartName="/xl/sharedStrings.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.'
                'spreadsheetml.sharedStrings+xml"/>'
            )
            ct_xml = ct_xml.replace("</Types>", f"{override}</Types>")
            members[ct_path] = ct_xml.encode("utf-8")

        # 4. Register the relationship from the workbook to sharedStrings.
        rels_path = "xl/_rels/workbook.xml.rels"
        rels_xml = members[rels_path].decode("utf-8")
        if "sharedStrings.xml" not in rels_xml:
            existing_ids = {
                int(m) for m in re.findall(r'Id="rId(\d+)"', rels_xml)
            }
            next_id = max(existing_ids, default=0) + 1
            new_rel = (
                f'<Relationship Id="rId{next_id}" '
                'Type="http://schemas.openxmlformats.org/officeDocument/'
                '2006/relationships/sharedStrings" '
                'Target="sharedStrings.xml"/>'
            )
            rels_xml = rels_xml.replace(
                "</Relationships>", f"{new_rel}</Relationships>"
            )
            members[rels_path] = rels_xml.encode("utf-8")

    # 5. Write everything back atomically.
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    try:
        with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in members.items():
                zf.writestr(name, data)
        shutil.move(str(tmp), xlsx_path)
    finally:
        if tmp.exists():
            tmp.unlink()


def _text_xml(text: str) -> str:
    """Re-escape text for embedding in <t> in sharedStrings.xml."""
    return xml_escape(text)
