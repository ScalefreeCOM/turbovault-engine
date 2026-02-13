"""
Excel Import Service for TurboVault Engine (SQLite-based).

This is a drop-in replacement for excel_import.py that uses openpyxl + an
in-memory SQLite database as the intermediate data layer instead of Pandas
DataFrames.  All Django ORM writes are identical to the original service.
"""

from __future__ import annotations

import logging
import sqlite3
from datetime import datetime
from typing import Any

import openpyxl
from django.db import transaction
from django.utils import timezone

from engine.models.hubs import Hub, HubColumn, HubSourceMapping
from engine.models.links import (
    Link,
    LinkColumn,
    LinkHubReference,
    LinkHubSourceMapping,
    LinkSourceMapping,
)
from engine.models.pit import PIT
from engine.models.prejoin import PrejoinDefinition, PrejoinExtractionColumn
from engine.models.project import Project
from engine.models.reference_table import (
    ReferenceTable,
    ReferenceTableSatelliteAssignment,
)
from engine.models.satellites import Satellite, SatelliteColumn
from engine.models.snapshot_control import SnapshotControlLogic, SnapshotControlTable
from engine.models.source_metadata import SourceColumn, SourceSystem, SourceTable

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_empty(val: Any) -> bool:
    """Return True if *val* is None or a stringified null-marker."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ("", "nan", "none")


def _clean(val: Any) -> str | None:
    """Return a stripped string or None."""
    if _is_empty(val):
        return None
    return str(val).strip()


def _rows(conn: sqlite3.Connection, table: str) -> list[sqlite3.Row]:
    """Return all rows from *table* as sqlite3.Row objects."""
    cur = conn.execute(f"SELECT * FROM [{table}]")
    return cur.fetchall()


def _row_get(row: sqlite3.Row, col: str) -> Any:
    """Safely get a column value from a Row; returns None if missing."""
    try:
        return row[col]
    except (IndexError, KeyError):
        return None


# ---------------------------------------------------------------------------
# Service
# ---------------------------------------------------------------------------

class ExcelImportSqliteService:
    """
    Service to import metadata from the legacy Excel format.

    Uses openpyxl to read the workbook and an in-memory SQLite database as
    the staging area for sheet data.  All Django ORM writes are identical
    to :class:`ExcelImportService`.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

        # Open workbook (read-only, data-only for perf)
        self._wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        self._sheet_names: list[str] = self._wb.sheetnames

        # In-memory SQLite staging DB
        self._conn = sqlite3.connect(":memory:")
        self._conn.row_factory = sqlite3.Row
        self._loaded_tables: set[str] = set()

        # Preload every sheet into SQLite
        for sheet_name in self._sheet_names:
            self._load_sheet_to_sqlite(sheet_name)

        # Internal caches – identical to the Pandas version
        self.project: Project | None = None
        self._source_systems: dict[str, SourceSystem] = {}
        self._source_tables: dict[str, SourceTable] = {}
        self._source_columns: dict[str, SourceColumn] = {}
        self._hubs: dict[str, Hub] = {}
        self._links: dict[str, Link] = {}
        self._satellites: dict[str, Satellite] = {}
        self._hub_columns: dict[str, HubColumn] = {}
        self._prejoins: dict[str, PrejoinDefinition] = {}
        self._extractions: dict[str, PrejoinExtractionColumn] = {}
        self._snapshot_control: SnapshotControlTable | None = None
        self._snapshot_logic: SnapshotControlLogic | None = None

    # ------------------------------------------------------------------
    # Sheet → SQLite loader
    # ------------------------------------------------------------------

    def _load_sheet_to_sqlite(self, sheet_name: str) -> None:
        """Read an openpyxl worksheet and insert all rows into a SQLite table."""
        ws = self._wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            return  # empty sheet

        columns = [str(c).strip().lower() for c in header_raw if c is not None]
        if not columns:
            return

        # Sanitised table name (sheet names may contain spaces)
        tbl = sheet_name  # we quote it in SQL with []

        col_defs = ", ".join(f"[{c}] TEXT" for c in columns)
        self._conn.execute(f"CREATE TABLE IF NOT EXISTS [{tbl}] ({col_defs})")

        placeholders = ", ".join("?" for _ in columns)
        insert_sql = f"INSERT INTO [{tbl}] ({', '.join(f'[{c}]' for c in columns)}) VALUES ({placeholders})"

        num_cols = len(columns)
        for row in rows_iter:
            # Trim row to header length (openpyxl can yield extra Nones)
            vals = list(row[:num_cols])
            # Convert non-None values to strings to match Pandas' text behaviour
            vals = [str(v).strip() if v is not None else None for v in vals]
            # Replace stringified null markers with actual None
            vals = [None if v is not None and v.strip().lower() in ("none", "") else v for v in vals]
            self._conn.execute(insert_sql, vals)

        self._conn.commit()
        self._loaded_tables.add(sheet_name)

    def _has_sheet(self, name: str) -> bool:
        return name in self._loaded_tables

    # ------------------------------------------------------------------
    # Public API (identical signature to ExcelImportService)
    # ------------------------------------------------------------------

    @transaction.atomic
    def import_metadata(
        self,
        project_name: str | None = None,
        description: str | None = None,
        project: Project | None = None,
        skip_snapshots: bool = False,
    ) -> Project:
        """Main entry point for importing metadata from Excel."""
        logger.info(f"Starting import from {self.file_path}")

        # 1. Create or Use Project
        if project:
            self.project = project
        else:
            if not project_name:
                raise ValueError(
                    "project_name is required if no project instance is provided"
                )
            self.project = Project.objects.create(
                name=project_name,
                description=description or f"Imported from {self.file_path}",
                config={},
            )

        # 2. Process Source Data
        if self._has_sheet("source_data"):
            self._process_source_data()

        # 3. Collect all Source Columns
        self._collect_all_source_columns()

        # 4. Process Hubs
        if self._has_sheet("standard_hub"):
            self._process_standard_hubs()
        if self._has_sheet("ref_hub"):
            self._process_reference_hubs()

        # 5. Process Prejoins (must happen before Links)
        if self._has_sheet("standard_link"):
            self._process_prejoins("standard_link")
        if self._has_sheet("non_historized_link"):
            self._process_prejoins("non_historized_link")

        # 5.1 Process Links
        if self._has_sheet("standard_link"):
            self._process_standard_links()
        if self._has_sheet("non_historized_link"):
            self._process_non_historized_links()

        # 6. Process Satellites
        sat_sheets = [
            "standard_satellite",
            "ref_sat",
            "non_historized_satellite",
            "multiactive_satellite",
        ]
        for sheet in sat_sheets:
            if self._has_sheet(sheet):
                self._process_satellites(sheet)

        # 7. Create default Snapshot Control
        self._create_default_snapshot_control(skip_creation=skip_snapshots)

        # 9. Process Reference Tables
        if self._has_sheet("ref_table"):
            self._process_reference_tables()

        # 10. Process PITs
        if self._has_sheet("pit"):
            self._process_pits()

        logger.info(f"Import completed. Project ID: {self.project.project_id}")
        return self.project

    # ------------------------------------------------------------------
    # Internal processing methods
    # ------------------------------------------------------------------

    def _process_source_data(self) -> None:
        """Processes the source_data sheet."""
        for row in _rows(self._conn, "source_data"):
            system_name = _clean(row["source_system"]) or "Unknown System"
            schema_name = _clean(row["source_schema_physical_name"]) or "public"

            key = f"{system_name}|{schema_name}"
            if key not in self._source_systems:
                system, _ = SourceSystem.objects.get_or_create(
                    project=self.project,
                    schema_name=schema_name,
                    name=system_name,
                    defaults={"database_name": None},
                )
                self._source_systems[key] = system

            system = self._source_systems[key]

            table_name = _clean(row["source_table_physical_name"])
            table_id = _clean(row["source_table_identifier"])

            if table_name and table_id not in self._source_tables:
                table = SourceTable.objects.create(
                    project=self.project,
                    source_system=system,
                    physical_table_name=table_name,
                    alias="",
                    record_source_value=_clean(
                        _row_get(row, "record_source_column")
                    )
                    or "",
                    static_part_of_record_source=_clean(
                        _row_get(row, "static_part_of_record_source_column")
                    )
                    or "",
                    load_date_value=_clean(_row_get(row, "load_date_column"))
                    or "sysdate()",
                )
                self._source_tables[table_id] = table
                if table_name not in self._source_tables:
                    self._source_tables[table_name] = table

    def _collect_all_source_columns(self) -> None:
        """Scans all sheets to find all mentioned source columns and creates them."""
        logger.info("Collecting source columns from all sheets...")
        columns_to_create: dict[str, set[str]] = {}

        sheets_to_scan = [
            ("standard_hub", "source_table_identifier", ["source_column_physical_name"]),
            ("ref_hub", "source_table_identifier", ["source_column_physical_name"]),
            ("standard_link", "source_table_identifier", ["source_column_physical_name"]),
            ("non_historized_link", "source_table_identifier", ["source_column_physical_name"]),
            ("standard_satellite", "source_table_identifier", ["source_column_physical_name"]),
            ("ref_sat", "source_table_identifier", ["source_column_physical_name"]),
            ("non_historized_satellite", "source_table_identifier", ["source_column_physical_name"]),
            ("multiactive_satellite", "source_table_identifier", ["source_column_physical_name", "multi_active_attributes"]),
            ("non_historized_link", "source_table_identifier", ["source_column_physical_name"]),
            ("non_historized_link", "prejoin_table_identifier", ["prejoin_table_column_name", "prejoin_extraction_column_name"]),
            ("standard_link", "source_table_identifier", ["source_column_physical_name"]),
            ("standard_link", "prejoin_table_identifier", ["prejoin_table_column_name", "prejoin_extraction_column_name"]),
        ]

        for sheet_name, table_col, data_cols in sheets_to_scan:
            if not self._has_sheet(sheet_name):
                continue

            for row in _rows(self._conn, sheet_name):
                table_key = _row_get(row, table_col)
                if _is_empty(table_key):
                    continue

                table_key = str(table_key)
                if table_key not in columns_to_create:
                    columns_to_create[table_key] = set()

                for dc in data_cols:
                    val = _row_get(row, dc)
                    if not _is_empty(val):
                        if dc == "multi_active_attributes":
                            for sub_val in str(val).split(";"):
                                columns_to_create[table_key].add(sub_val.strip())
                        else:
                            columns_to_create[table_key].add(str(val).strip())

        # Create columns in the DB
        for table_key, cols in columns_to_create.items():
            table = self._source_tables.get(table_key)
            if not table:
                logger.warning(
                    f"Source table {table_key} not found in source_data, "
                    "but referenced columns exist."
                )
                continue

            for col_name in cols:
                column, _ = SourceColumn.objects.get_or_create(
                    source_table=table,
                    source_column_physical_name=col_name,
                    defaults={"source_column_datatype": ""},
                )
                self._source_columns[f"{table_key}|{col_name}"] = column
                phys_name = table.physical_table_name
                self._source_columns[f"{phys_name}|{col_name}"] = column

    def _process_standard_hubs(self) -> None:
        """Processes standard_hub sheet."""
        for row in _rows(self._conn, "standard_hub"):
            hub_name = _clean(row["target_hub_table_physical_name"])
            if not hub_name:
                continue

            if hub_name not in self._hubs:
                hub = Hub.objects.create(
                    project=self.project,
                    hub_physical_name=hub_name,
                    hub_type=Hub.HubType.STANDARD,
                    hub_hashkey_name=_clean(row["target_primary_key_physical_name"]),
                    create_record_tracking_satellite=(
                        str(_row_get(row, "record_tracking_satellite")).upper() == "TRUE"
                    ),
                    create_effectivity_satellite=False,
                )
                self._hubs[hub_name] = hub
                hub_id = _clean(row["hub_identifier"])
                if hub_id:
                    self._hubs[hub_id] = hub

            hub = self._hubs[hub_name]

            col_name = _clean(_row_get(row, "business_key_physical_name")) or _clean(
                _row_get(row, "source_column_physical_name")
            )
            if not col_name:
                continue

            hub_column, _ = HubColumn.objects.get_or_create(
                hub=hub,
                column_name=col_name,
                defaults={"column_type": HubColumn.ColumnType.BUSINESS_KEY},
            )

            source_table_key = _clean(row["source_table_identifier"])
            source_col_name = _clean(row["source_column_physical_name"])
            source_col = self._source_columns.get(
                f"{source_table_key}|{source_col_name}"
            )

            if source_col:
                HubSourceMapping.objects.get_or_create(
                    hub_column=hub_column,
                    source_column=source_col,
                    defaults={
                        "is_primary_source": str(
                            _row_get(row, "is_primary_source")
                        ).upper()
                        == "TRUE"
                    },
                )

        # Post-processing: ensure at least one primary source per hub column
        for hub in Hub.objects.filter(project=self.project):
            for col in hub.columns.all():
                if not HubSourceMapping.objects.filter(
                    hub_column=col, is_primary_source=True
                ).exists():
                    first_mapping = HubSourceMapping.objects.filter(
                        hub_column=col
                    ).first()
                    if first_mapping:
                        first_mapping.is_primary_source = True
                        first_mapping.save()

    def _process_reference_hubs(self) -> None:
        """Processes ref_hub sheet."""
        for row in _rows(self._conn, "ref_hub"):
            hub_name = _clean(row["target_reference_table_physical_name"])
            if not hub_name:
                continue

            if hub_name not in self._hubs:
                hub = Hub.objects.create(
                    project=self.project,
                    hub_physical_name=hub_name,
                    hub_type=Hub.HubType.REFERENCE,
                    hub_hashkey_name=None,
                    create_record_tracking_satellite=False,
                    create_effectivity_satellite=False,
                )
                self._hubs[hub_name] = hub
                hub_id = _clean(_row_get(row, "reference_hub_identifier"))
                if hub_id:
                    self._hubs[hub_id] = hub

            hub = self._hubs[hub_name]

            source_col_name = _clean(row["source_column_physical_name"])
            if not source_col_name:
                continue

            hub_column, _ = HubColumn.objects.get_or_create(
                hub=hub,
                column_name=source_col_name,
                defaults={"column_type": HubColumn.ColumnType.REFERENCE_KEY},
            )

            source_table_key = _clean(row["source_table_identifier"])
            source_col = self._source_columns.get(
                f"{source_table_key}|{source_col_name}"
            )

            if source_col:
                HubSourceMapping.objects.get_or_create(
                    hub_column=hub_column,
                    source_column=source_col,
                    defaults={"is_primary_source": True},
                )

    def _process_standard_links(self) -> None:
        """Processes standard_link sheet."""
        self._process_links_generic("standard_link", Link.LinkType.STANDARD, "link_identifier")

    def _process_non_historized_links(self) -> None:
        """Processes non_historized_link sheet."""
        self._process_links_generic(
            "non_historized_link", Link.LinkType.NON_HISTORIZED, "nh_link_identifier"
        )

    def _process_links_generic(
        self,
        sheet: str,
        link_type: str,
        id_col: str,
    ) -> None:
        """
        Shared logic for standard and non-historized link sheets.

        Replicates the Pandas version's forward-fill on
        ``target_link_table_physical_name``, groupby on link name, and
        hub-ref vs payload separation.
        """
        all_rows = _rows(self._conn, sheet)
        if not all_rows:
            return

        # Check if columns exist
        col_names = all_rows[0].keys()
        has_hub_id = "hub_identifier" in col_names

        # Forward-fill target_link_table_physical_name (replicates df.ffill())
        rows_as_dicts: list[dict[str, Any]] = [dict(r) for r in all_rows]
        last_link_name: str | None = None
        for rd in rows_as_dicts:
            if not _is_empty(rd.get("target_link_table_physical_name")):
                last_link_name = rd["target_link_table_physical_name"]
            else:
                rd["target_link_table_physical_name"] = last_link_name

        # Group by link name
        link_groups: dict[str, list[dict[str, Any]]] = {}
        for rd in rows_as_dicts:
            ln = _clean(rd.get("target_link_table_physical_name"))
            if not ln:
                continue
            link_groups.setdefault(ln, []).append(rd)

        for link_name, link_rows in link_groups.items():
            # Create Link
            row_sample = link_rows[0]
            if link_name not in self._links:
                link = Link.objects.create(
                    project=self.project,
                    link_physical_name=link_name,
                    link_hashkey_name=_clean(
                        row_sample.get("target_primary_key_physical_name")
                    )
                    or f"lk_{link_name}",
                    link_type=link_type,
                )
                self._links[link_name] = link
                lid = _clean(row_sample.get(id_col))
                if lid:
                    self._links[lid] = link

            link = self._links[link_name]

            # Separate hub-ref rows from payload rows
            if not has_hub_id:
                ref_rows: list[dict[str, Any]] = []
                payload_rows = link_rows
            else:
                ref_rows = [
                    r for r in link_rows if not _is_empty(r.get("hub_identifier"))
                ]
                payload_rows = [
                    r for r in link_rows if _is_empty(r.get("hub_identifier"))
                ]

            # === Process Hub References ===
            if ref_rows:
                # Group by hub_identifier
                hub_groups: dict[str, list[dict[str, Any]]] = {}
                for r in ref_rows:
                    hid = str(r["hub_identifier"])
                    hub_groups.setdefault(hid, []).append(r)

                for hub_id_str, hub_rows in hub_groups.items():
                    hub = self._hubs.get(hub_id_str)
                    if not hub:
                        hub = Hub.objects.filter(
                            project=self.project, hub_physical_name=hub_id_str
                        ).first()

                    if not hub:
                        logger.warning(
                            f"Hub {hub_id_str} not found in Link {link_name}"
                        )
                        continue

                    # Fill NULLs in target_column_physical_name
                    for r in hub_rows:
                        if _is_empty(r.get("target_column_physical_name")):
                            r["target_column_physical_name"] = ""

                    # Group by alias
                    alias_groups: dict[str, list[dict[str, Any]]] = {}
                    for r in hub_rows:
                        alias_groups.setdefault(
                            r["target_column_physical_name"], []
                        ).append(r)

                    for alias, alias_rows in alias_groups.items():
                        hk_col = _clean(
                            alias_rows[0].get("target_primary_key_physical_name")
                        )
                        final_alias = alias if alias != hk_col else ""
                        sort_order = int(
                            alias_rows[0].get("target_column_sort_order") or 0
                        )

                        lhr = LinkHubReference.objects.create(
                            link=link,
                            hub=hub,
                            hub_hashkey_alias_in_link=final_alias,
                            sort_order=sort_order,
                        )

                        # Sort by target_column_sort_order
                        alias_rows_sorted = sorted(
                            alias_rows,
                            key=lambda r: int(r.get("target_column_sort_order") or 0),
                        )

                        hub_cols = list(
                            hub.columns.filter(
                                column_type=HubColumn.ColumnType.BUSINESS_KEY
                            ).order_by("sort_order")
                        )
                        if not hub_cols:
                            hub_cols = list(
                                hub.columns.filter(
                                    column_type=HubColumn.ColumnType.REFERENCE_KEY
                                ).order_by("sort_order")
                            )

                        for idx, r in enumerate(alias_rows_sorted):
                            if idx >= len(hub_cols):
                                continue

                            hub_col = hub_cols[idx]

                            source_table_id = _clean(r.get("source_table_identifier"))
                            source_col_name = _clean(
                                r.get("source_column_physical_name")
                            )
                            prejoin_alias = _clean(
                                r.get("prejoin_target_column_alias")
                            )
                            ext_col_name = _clean(
                                r.get("prejoin_extraction_column_name")
                            )

                            prejoin_ext = None
                            source_col = None

                            if prejoin_alias or ext_col_name:
                                key_candidate = (
                                    prejoin_alias or ext_col_name or source_col_name
                                )
                                extraction_key = f"{link_name}|{key_candidate}"
                                prejoin_ext = self._extractions.get(extraction_key)

                            if not prejoin_ext and source_table_id and source_col_name:
                                source_col = self._source_columns.get(
                                    f"{source_table_id}|{source_col_name}"
                                )

                            if source_col or prejoin_ext:
                                LinkHubSourceMapping.objects.create(
                                    link_hub_reference=lhr,
                                    standard_hub_column=hub_col,
                                    source_column=source_col,
                                    prejoin_extraction_column=prejoin_ext,
                                )

            # === Process Payload ===
            for r in payload_rows:
                col_name = (
                    _clean(r.get("target_column_physical_name"))
                    or _clean(r.get("prejoin_target_column_alias"))
                    or _clean(r.get("prejoin_extraction_column_name"))
                    or _clean(r.get("source_column_physical_name"))
                )
                if not col_name:
                    continue

                hk_def = r.get("target_primary_key_physical_name")
                is_key = bool(hk_def and not _is_empty(hk_def))
                col_type = (
                    LinkColumn.ColumnType.DEPENDANT_CHILD_KEY
                    if is_key
                    else LinkColumn.ColumnType.PAYLOAD
                )

                lc, _ = LinkColumn.objects.get_or_create(
                    link=link,
                    column_name=col_name,
                    defaults={
                        "column_type": col_type,
                        "sort_order": int(r.get("target_column_sort_order") or 0),
                    },
                )

                prejoin_alias = _clean(r.get("prejoin_target_column_alias"))
                ext_col_name = _clean(r.get("prejoin_extraction_column_name"))
                prejoin_ext = None
                source_col = None

                if prejoin_alias or ext_col_name:
                    key_candidate = prejoin_alias or ext_col_name or col_name
                    extraction_key = f"{link_name}|{key_candidate}"
                    prejoin_ext = self._extractions.get(extraction_key)

                if prejoin_ext:
                    source_col = prejoin_ext.source_column
                elif not source_col:
                    source_table_id = _clean(r.get("source_table_identifier"))
                    source_col_name = _clean(r.get("source_column_physical_name"))
                    if source_table_id and source_col_name:
                        source_col = self._source_columns.get(
                            f"{source_table_id}|{source_col_name}"
                        )

                if source_col:
                    LinkSourceMapping.objects.get_or_create(
                        link_column=lc,
                        source_column=source_col,
                    )

    def _process_satellites(self, sheet_type: str) -> None:
        """Processes all types of satellite sheets."""
        for row in _rows(self._conn, sheet_type):
            sat_name_col = (
                "target_reference_table_physical_name"
                if sheet_type == "ref_sat"
                else "target_satellite_table_physical_name"
            )
            sat_name = _clean(_row_get(row, sat_name_col))
            if not sat_name:
                continue

            if sat_name not in self._satellites:
                sat_type_map = {
                    "standard_satellite": Satellite.SatelliteType.STANDARD,
                    "ref_sat": Satellite.SatelliteType.REFERENCE,
                    "non_historized_satellite": Satellite.SatelliteType.NON_HISTORIZED,
                    "multiactive_satellite": Satellite.SatelliteType.MULTI_ACTIVE,
                }

                parent_id = (
                    _clean(_row_get(row, "parent_identifier"))
                    or _clean(_row_get(row, "parent_table_identifier"))
                    or _clean(_row_get(row, "nh_link_identifier"))
                    or _clean(_row_get(row, "referenced_hub"))
                    or _clean(_row_get(row, "parent_hub"))
                )

                parent_hub = self._hubs.get(parent_id)
                parent_link = self._links.get(parent_id)

                if not parent_hub and not parent_link:
                    parent_hub = Hub.objects.filter(
                        project=self.project, hub_physical_name=parent_id
                    ).first()
                    parent_link = Link.objects.filter(
                        project=self.project, link_physical_name=parent_id
                    ).first()

                source_table_key = _clean(_row_get(row, "source_table_identifier"))
                source_table = self._source_tables.get(source_table_key)

                if (parent_hub or parent_link) and source_table:
                    satellite = Satellite.objects.create(
                        project=self.project,
                        satellite_physical_name=sat_name,
                        satellite_type=sat_type_map[sheet_type],
                        parent_hub=parent_hub,
                        parent_link=parent_link,
                        source_table=source_table,
                    )
                    self._satellites[sat_name] = satellite
                    sat_id_col = (
                        "ma_satellite_identifier"
                        if sheet_type == "multiactive_satellite"
                        else (
                            "reference_satellite_identifier"
                            if sheet_type == "ref_sat"
                            else "satellite_identifier"
                        )
                    )
                    sat_id = _clean(_row_get(row, sat_id_col))
                    if sat_id:
                        self._satellites[sat_id] = satellite
                else:
                    logger.warning(
                        f"Could not find parent {parent_id} or source table "
                        f"{source_table_key} for satellite {sat_name}"
                    )
                    continue

            satellite = self._satellites[sat_name]

            # Satellite Columns
            cols_to_add: list[tuple[str, bool]] = []

            regular_col = _clean(_row_get(row, "source_column_physical_name"))
            if regular_col:
                cols_to_add.append((regular_col, False))

            if sheet_type == "multiactive_satellite":
                ma_attrs = _clean(_row_get(row, "multi_active_attributes"))
                if ma_attrs:
                    for s in str(ma_attrs).split(";"):
                        cols_to_add.append((s.strip(), True))

            for scn, is_ma_key in cols_to_add:
                source_table_key = _clean(_row_get(row, "source_table_identifier"))
                source_col = self._source_columns.get(f"{source_table_key}|{scn}")
                if source_col:
                    target_col_name = scn
                    if not is_ma_key:
                        tcn = _clean(_row_get(row, "target_column_physical_name"))
                        if tcn:
                            target_col_name = tcn

                    SatelliteColumn.objects.get_or_create(
                        satellite=satellite,
                        source_column=source_col,
                        defaults={
                            "is_multi_active_key": is_ma_key,
                            "include_in_delta_detection": sheet_type
                            != "non_historized_satellite",
                            "target_column_name": target_col_name,
                        },
                    )

    def _process_prejoins(self, sheet: str) -> None:
        """Processes prejoin definitions and extractions from a link sheet."""
        all_rows = _rows(self._conn, sheet)
        if not all_rows:
            return

        rows_as_dicts: list[dict[str, Any]] = [dict(r) for r in all_rows]

        # Forward fill prejoin_table_identifier within each
        # (target_link_table_physical_name, source_table_identifier) group —
        # identical to the Pandas grouped ffill.
        col_names_set = set(rows_as_dicts[0].keys()) if rows_as_dicts else set()
        group_cols = ["target_link_table_physical_name", "source_table_identifier"]
        prejoin_cols = ["prejoin_table_identifier"]

        active_group_cols = [c for c in group_cols if c in col_names_set]
        active_prejoin_cols = [c for c in prejoin_cols if c in col_names_set]

        if active_group_cols and active_prejoin_cols:
            # Track the last non-null value per group per prejoin column
            last_vals: dict[tuple[str, ...], dict[str, str | None]] = {}
            for rd in rows_as_dicts:
                grp_key = tuple(str(rd.get(c) or "") for c in active_group_cols)
                if grp_key not in last_vals:
                    last_vals[grp_key] = {pc: None for pc in active_prejoin_cols}
                for pc in active_prejoin_cols:
                    if not _is_empty(rd.get(pc)):
                        last_vals[grp_key][pc] = rd[pc]
                    else:
                        rd[pc] = last_vals[grp_key][pc]

        def get_table_and_id(
            tid: str | None,
        ) -> tuple[SourceTable | None, str | None]:
            if not tid:
                return None, None
            if tid in self._source_tables:
                return self._source_tables[tid], tid
            for k, v in self._source_tables.items():
                if k.lower() == str(tid).lower():
                    return v, k
            return None, None

        for rd in rows_as_dicts:
            source_table_id = _clean(rd.get("source_table_identifier"))
            target_table_id = _clean(rd.get("prejoin_table_identifier"))

            if not target_table_id:
                continue

            link_name = _clean(rd.get("target_link_table_physical_name"))

            source_table, source_table_id = get_table_and_id(source_table_id)
            target_table, target_table_id = get_table_and_id(target_table_id)

            if not source_table or not target_table:
                continue

            prejoin_key = f"{source_table_id}|{target_table_id}"
            if prejoin_key not in self._prejoins:
                prejoin = PrejoinDefinition.objects.create(
                    project=self.project,
                    source_table=source_table,
                    prejoin_target_table=target_table,
                    prejoin_operator=PrejoinDefinition.JoinOperator.AND,
                )

                source_col_name = _clean(rd.get("source_column_physical_name"))
                target_col_name = _clean(rd.get("prejoin_table_column_name"))

                source_col = self._source_columns.get(
                    f"{source_table_id}|{source_col_name}"
                )
                target_col = self._source_columns.get(
                    f"{target_table_id}|{target_col_name}"
                )

                if source_col:
                    prejoin.prejoin_condition_source_column.add(source_col)
                if target_col:
                    prejoin.prejoin_condition_target_column.add(target_col)

                self._prejoins[prejoin_key] = prejoin

            prejoin = self._prejoins.get(prejoin_key)
            if not prejoin:
                continue

            extraction_col_name = _clean(rd.get("prejoin_extraction_column_name"))
            if extraction_col_name:
                source_col = self._source_columns.get(
                    f"{target_table_id}|{extraction_col_name}"
                )

                if source_col:
                    alias = _clean(rd.get("prejoin_target_column_alias"))
                    extraction, _ = PrejoinExtractionColumn.objects.update_or_create(
                        prejoin=prejoin,
                        source_column=source_col,
                        defaults={"prejoin_target_column_alias": alias},
                    )
                    self._extractions[
                        f"{link_name}|{alias or extraction_col_name}"
                    ] = extraction

    def _process_reference_tables(self) -> None:
        """Processes ref_table sheet."""
        for row in _rows(self._conn, "ref_table"):
            ref_table_name = _clean(
                _row_get(row, "target_reference_table_physical_name")
            )
            hub_id = _clean(_row_get(row, "referenced_hub"))

            if not ref_table_name:
                continue

            hub = self._hubs.get(hub_id)
            if not hub:
                hub = Hub.objects.filter(
                    project=self.project, hub_physical_name=hub_id
                ).first()

            if not hub:
                logger.warning(
                    f"Hub {hub_id} not found for reference table {ref_table_name}"
                )
                continue

            hist_type_map = {
                "TRUE": ReferenceTable.HistorizationType.FULL,
                "FALSE": ReferenceTable.HistorizationType.LATEST,
                "FULL": ReferenceTable.HistorizationType.FULL,
                "LATEST": ReferenceTable.HistorizationType.LATEST,
            }
            hist_val = _clean(_row_get(row, "historized")) or "LATEST"

            ref_table, _ = ReferenceTable.objects.get_or_create(
                project=self.project,
                reference_table_physical_name=ref_table_name,
                defaults={
                    "reference_hub": hub,
                    "historization_type": hist_type_map.get(
                        hist_val.upper(), ReferenceTable.HistorizationType.LATEST
                    ),
                },
            )

            sat_id = _clean(_row_get(row, "referenced_satellite"))
            sat = self._satellites.get(sat_id)
            if not sat:
                sat = Satellite.objects.filter(
                    project=self.project, satellite_physical_name=sat_id
                ).first()

            if sat:
                assignment, _ = (
                    ReferenceTableSatelliteAssignment.objects.get_or_create(
                        reference_table=ref_table, reference_satellite=sat
                    )
                )

                include_str = _clean(_row_get(row, "included_columns"))
                exclude_str = _clean(_row_get(row, "excluded_columns"))

                if include_str:
                    for col_name in include_str.split(","):
                        col = SatelliteColumn.objects.filter(
                            satellite=sat, target_column_name=col_name.strip()
                        ).first()
                        if col:
                            assignment.include_columns.add(col)

                if exclude_str:
                    for col_name in exclude_str.split(","):
                        col = SatelliteColumn.objects.filter(
                            satellite=sat, target_column_name=col_name.strip()
                        ).first()
                        if col:
                            assignment.exclude_columns.add(col)

    def _process_pits(self) -> None:
        """Processes pit sheet."""
        for row in _rows(self._conn, "pit"):
            pit_name = _clean(_row_get(row, "pit_physical_table_name"))
            entity_id = _clean(_row_get(row, "tracked_entity"))

            if not pit_name:
                continue

            hub = self._hubs.get(entity_id)
            link = self._links.get(entity_id)

            if not hub and not link:
                hub = Hub.objects.filter(
                    project=self.project, hub_physical_name=entity_id
                ).first()
                link = Link.objects.filter(
                    project=self.project, link_physical_name=entity_id
                ).first()

            if not hub and not link:
                logger.warning(
                    f"Tracked entity {entity_id} not found for PIT {pit_name}"
                )
                continue

            pit = PIT.objects.create(
                project=self.project,
                pit_physical_name=pit_name,
                tracked_entity_type=(
                    PIT.TrackedEntityType.HUB if hub else PIT.TrackedEntityType.LINK
                ),
                tracked_hub=hub,
                tracked_link=link,
                snapshot_control_table=self._snapshot_control,
                snapshot_control_logic=self._snapshot_logic,
                dimension_key_column_name=_clean(
                    _row_get(row, "dimension_key_name")
                ),
                pit_type=_clean(_row_get(row, "pit_type")),
                custom_record_source=_clean(
                    _row_get(row, "custom_record_source")
                ),
            )

            sat_ids = _row_get(row, "satellite_identifiers")
            if not _is_empty(sat_ids):
                for sid in str(sat_ids).replace(";", ",").split(","):
                    sid = sid.strip()
                    if not sid:
                        continue
                    sat = self._satellites.get(sid)
                    if not sat:
                        sat = Satellite.objects.filter(
                            project=self.project, satellite_physical_name=sid
                        ).first()
                    if not sat:
                        sat = Satellite.objects.filter(
                            project=self.project,
                            satellite_physical_name__iexact=sid,
                        ).first()
                    if sat:
                        pit.satellites.add(sat)

    def _create_default_snapshot_control(
        self, skip_creation: bool = False
    ) -> None:
        """Creates a default snapshot control table and logic rule."""
        if self._snapshot_control:
            return

        existing_table = SnapshotControlTable.objects.filter(
            project=self.project
        ).first()
        if existing_table:
            self._snapshot_control = existing_table
            self._snapshot_logic = SnapshotControlLogic.objects.filter(
                snapshot_control_table=existing_table
            ).first()
            return

        if skip_creation:
            return

        today = timezone.now().date()
        self._snapshot_control = SnapshotControlTable.objects.create(
            project=self.project,
            snapshot_start_date=datetime(today.year - 5, 1, 1).date(),
            snapshot_end_date=datetime(today.year + 5, 12, 31).date(),
            daily_snapshot_time=timezone.make_aware(
                datetime(2000, 1, 1, 8, 0, 0)
            ).time(),
        )

        self._snapshot_logic = SnapshotControlLogic.objects.create(
            snapshot_control_table=self._snapshot_control,
            snapshot_control_logic_column_name="is_active",
            snapshot_component=SnapshotControlLogic.SnapshotComponent.BEGINNING_OF_MONTH,
            snapshot_duration=1,
            snapshot_unit=SnapshotControlLogic.SnapshotUnit.YEAR,
            snapshot_forever=False,
        )

    def _get_val(self, row: sqlite3.Row | dict[str, Any], col: str) -> str | None:
        """Helper to get a cleaned string value or None if empty."""
        if isinstance(row, dict):
            val = row.get(col)
        else:
            val = _row_get(row, col)
        return _clean(val)

    def close(self) -> None:
        """Close the in-memory SQLite connection and workbook."""
        self._conn.close()
        self._wb.close()
